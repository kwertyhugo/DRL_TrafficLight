"""
run_Balibago_SP_DDPG_batch_test.py

Runs all 30 Balibago demand_test route files through the DDPG SIGNALIZED
simulation in test mode and writes:
    Balibago_traci/batch_results/SP_DDPG_Results.xlsx

Matches SignalizedPedestrianDDPG.py exactly:
  - signalizedPed.sumocfg,  stepLength=0.1,  no hard step cap
  - North (4902876117): 8-phase %8, state=12 (8q+4ph), norm /100,
    base={0:45, 2:130, 4:30, 6:90}, ACTION_SIZE=4, ACTION_SCALE=25
    NORTH_GREEN_PHASES={0:0, 2:1, 4:2, 6:3} — index into 4-action vector
  - South (12188714):   8-phase %8, state=9  (5q+4ph), norm /100,
    base={0:25, 2:30,  4:40, 6:45}, same ACTION_SIZE=4
    SOUTH_GREEN_PHASES={0:0, 2:1, 4:2, 6:3}
  - Pedestrian junction subscriptions on both intersections
  - Models: North1_DDPGAgent / South1_DDPGAgent
            loaded from ./Balibago_traci/output_DDPG/
  - Metric interval: every 60 s (step_counter % 600 == 0)

Metrics per run:
  - Mean Veh. Time Loss (s)   — <tripinfo timeLoss>
  - Mean Ped. Time Loss (s)   — <personinfo timeLoss>
  - Mean Queue Length (m)     — avg jam across 13 detectors
  - Throughput (veh/hr)       — vehicles cleared, normalised to per-hour
  - Mean North Queue (m)      — avg jam e2_0–e2_7
  - Mean South Queue (m)      — avg jam e2_8–e2_12
"""

import os
import sys
import xml.etree.ElementTree as ET
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
os.makedirs('./Balibago_traci/output_DDPG', exist_ok=True)

if 'SUMO_HOME' in os.environ:
    sys.path.append(os.path.join(os.environ['SUMO_HOME'], 'tools'))
else:
    sys.exit("Please declare environment variable 'SUMO_HOME'")

import traci
from keras.utils import to_categorical
from models.DDPG import DDPGAgent as ddpg

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
SUMOCFG        = 'Balibago_traci/signalizedPed.sumocfg'
DEMAND_DIR     = 'Balibago_traci/demand_test'
OUTPUT_DIR     = 'Balibago_traci/batch_results/SP_DDPG'
XLSX_OUT       = 'Balibago_traci/batch_results/SP_DDPG_Results.xlsx'
STEP_LENGTH    = 0.1
METRIC_STEPS   = 600      # every 60 s — matches original (step_counter % 600 == 0)
ACTION_SIZE    = 4
ACTION_SCALE   = 25.0
YELLOW_DUR     = 5.0

DETECTOR_IDS   = [f"e2_{i}" for i in range(13)]
DETECTOR_COUNT = 13
NORTH_IDS      = [f"e2_{i}" for i in range(8)]
SOUTH_IDS      = [f"e2_{i}" for i in range(8, 13)]

NORTH_TL           = "4902876117"
NORTH_PHASES       = 8
NORTH_BASE         = {0: 45, 2: 130, 4: 30, 6: 90}
NORTH_GREEN_PHASES = {0: 0, 2: 1, 4: 2, 6: 3}   # phase → action vector index

SOUTH_TL           = "12188714"
SOUTH_PHASES       = 8
SOUTH_BASE         = {0: 25, 2: 30, 4: 40, 6: 45}
SOUTH_GREEN_PHASES = {0: 0, 2: 1, 4: 2, 6: 3}

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(os.path.dirname(XLSX_OUT), exist_ok=True)

# ---------------------------------------------------------------------------
# Load agents (trainMode=0 → no noise)
# ---------------------------------------------------------------------------
action_low  = np.full(ACTION_SIZE, -1.0, dtype=np.float32)
action_high = np.full(ACTION_SIZE,  1.0, dtype=np.float32)

NorthAgent = ddpg(state_size=12, action_size=ACTION_SIZE,
                  action_low=action_low, action_high=action_high,
                  actor_lr=0.0001, critic_lr=0.001,
                  name='North1_DDPGAgent')
SouthAgent = ddpg(state_size=9,  action_size=ACTION_SIZE,
                  action_low=action_low, action_high=action_high,
                  actor_lr=0.0001, critic_lr=0.001,
                  name='South1_DDPGAgent')

model_path = './Balibago_traci/output_DDPG/'
if os.path.exists(model_path) and os.listdir(model_path):
    try:
        NorthAgent.load()
        SouthAgent.load()
        print(f"\n{'='*65}\n  North1 + South1 DDPG models loaded.\n{'='*65}")
    except Exception as e:
        print(f"  WARNING: Could not load models: {e}. Using fresh weights.")
else:
    print("  WARNING: No saved models found. Using fresh weights.")

# ---------------------------------------------------------------------------
# Test plan
# ---------------------------------------------------------------------------
TEST_PLAN = []
for scenario in ["normal", "slow", "jam"]:
    for i in range(1, 11):
        TEST_PLAN.append({
            "label":     f"{scenario.capitalize()} {i}",
            "scenario":  scenario,
            "rou_file":  os.path.join(DEMAND_DIR, f"flows_{scenario}_traffic_{i:02d}.rou.xml"),
            "trips_out": os.path.join(OUTPUT_DIR, f"SP_DDPG_trips_{scenario}_{i:02d}.xml"),
            "stats_out": os.path.join(OUTPUT_DIR, f"SP_DDPG_stats_{scenario}_{i:02d}.xml"),
        })

# ---------------------------------------------------------------------------
# Subscriptions
# ---------------------------------------------------------------------------

def subscribe_detectors():
    ctx_vars = [traci.constants.VAR_TYPE, traci.constants.VAR_WAITING_TIME]
    met_vars = [traci.constants.JAM_LENGTH_METERS, traci.constants.VAR_INTERVAL_NUMBER]
    for det in DETECTOR_IDS:
        traci.lanearea.subscribeContext(det, traci.constants.CMD_GET_VEHICLE_VARIABLE, 3, ctx_vars)
        traci.lanearea.subscribe(det, met_vars)

def subscribe_junctions():
    for jid in [NORTH_TL, SOUTH_TL]:
        traci.junction.subscribeContext(
            jid, traci.constants.CMD_GET_PERSON_VARIABLE,
            10.0, [traci.constants.VAR_WAITING_TIME])

# ---------------------------------------------------------------------------
# State helpers — exact match to SP DDPG
# ---------------------------------------------------------------------------

def weighted_waits(det_id):
    data = traci.lanearea.getContextSubscriptionResults(det_id)
    if not data: return 0
    wmap = {"car": 1.0, "jeep": 1.5, "bus": 2.2, "truck": 2.5,
            "motorcycle": 0.3, "tricycle": 0.5}
    return sum(d.get(traci.constants.VAR_WAITING_TIME, 0)
               * wmap.get(d.get(traci.constants.VAR_TYPE, "car"), 1.0)
               for d in data.values())

def get_north_state(phase):
    # 8 queues norm /100 + 4 phase one-hot = 12
    q  = np.array([weighted_waits(f"e2_{i}") for i in range(8)],
                  dtype=np.float32) / 100.0
    ph = to_categorical(phase // 2, num_classes=4).flatten()
    return np.concatenate([q, ph]).astype(np.float32)

def get_south_state(phase):
    # 5 queues norm /100 + 4 phase one-hot = 9
    q  = np.array([weighted_waits(f"e2_{i}") for i in range(8, 13)],
                  dtype=np.float32) / 100.0
    ph = to_categorical(phase // 2, num_classes=4).flatten()
    return np.concatenate([q, ph]).astype(np.float32)

def apply_north_phase(action_vec, current_phase):
    next_phase = (current_phase + 1) % NORTH_PHASES
    traci.trafficlight.setPhase(NORTH_TL, next_phase)
    if next_phase % 2 == 1:
        dur = YELLOW_DUR
    else:
        act_idx    = NORTH_GREEN_PHASES.get(next_phase, 0)
        adjustment = float(action_vec[act_idx]) * ACTION_SCALE
        dur        = max(5.0, min(180.0, NORTH_BASE.get(next_phase, 30) + adjustment))
    traci.trafficlight.setPhaseDuration(NORTH_TL, float(dur))
    return next_phase, float(dur)

def apply_south_phase(action_vec, current_phase):
    next_phase = (current_phase + 1) % SOUTH_PHASES
    traci.trafficlight.setPhase(SOUTH_TL, next_phase)
    if next_phase % 2 == 1:
        dur = YELLOW_DUR
    else:
        act_idx    = SOUTH_GREEN_PHASES.get(next_phase, 0)
        adjustment = float(action_vec[act_idx]) * ACTION_SCALE
        dur        = max(5.0, min(180.0, SOUTH_BASE.get(next_phase, 30) + adjustment))
    traci.trafficlight.setPhaseDuration(SOUTH_TL, float(dur))
    return next_phase, float(dur)

# ---------------------------------------------------------------------------
# Tripinfo parser
# ---------------------------------------------------------------------------

def parse_tripinfo(path):
    try:
        root = ET.parse(path).getroot()
    except (FileNotFoundError, ET.ParseError):
        return 0.0, 0.0
    veh, ped = [], []
    for tag in root:
        tl = tag.get("timeLoss")
        if tl is None: continue
        (veh if tag.tag == "tripinfo" else ped).append(float(tl))
    return (sum(veh) / len(veh) if veh else 0.0,
            sum(ped) / len(ped) if ped else 0.0)

# ---------------------------------------------------------------------------
# Single run
# ---------------------------------------------------------------------------

def run_simulation(test):
    sumo_cmd = [
        'sumo', '-c', SUMOCFG,
        '--route-files',        test["rou_file"],
        '--step-length',        str(STEP_LENGTH),
        '--delay',              '0',
        '--lateral-resolution', '0.1',
        '--statistic-output',   test["stats_out"],
        '--tripinfo-output',    test["trips_out"],
        '--no-warnings',        'true',
    ]
    print(f"\n{'='*65}\n  Running: {test['label']}\n{'='*65}")
    traci.start(sumo_cmd)
    subscribe_detectors()
    subscribe_junctions()

    n_phase, n_dur = 0, float(NORTH_BASE[0])
    s_phase, s_dur = 0, float(SOUTH_BASE[0])
    n_prev_act = np.zeros(ACTION_SIZE, dtype=np.float32)
    s_prev_act = np.zeros(ACTION_SIZE, dtype=np.float32)

    step = 0
    jam_total = north_jam_total = south_jam_total = 0.0
    tp_total  = 0
    obs       = 0

    while traci.simulation.getMinExpectedNumber() > 0:
        step += 1
        n_dur -= STEP_LENGTH
        s_dur -= STEP_LENGTH

        north_decision = (n_dur <= 0) and (n_phase % 2 == 0)
        south_decision = (s_dur <= 0) and (s_phase % 2 == 0)

        if north_decision:
            n_act = NorthAgent.get_action(get_north_state(n_phase), add_noise=False)
            n_prev_act = n_act
        elif n_dur <= 0:
            n_act = n_prev_act   # yellow — carry previous vector
        else:
            n_act = n_prev_act

        if south_decision:
            s_act = SouthAgent.get_action(get_south_state(s_phase), add_noise=False)
            s_prev_act = s_act
        elif s_dur <= 0:
            s_act = np.zeros(ACTION_SIZE, dtype=np.float32)  # yellow — no adjustment
        else:
            s_act = s_prev_act

        if n_dur <= 0:
            n_phase, n_dur = apply_north_phase(n_act, n_phase)
        if s_dur <= 0:
            s_phase, s_dur = apply_south_phase(s_act, s_phase)

        # Metric collection — matches original: step_counter % 600 == 0
        if step % METRIC_STEPS == 0:
            obs += 1
            jam = north_jam = south_jam = 0.0
            tp  = 0
            for det in DETECTOR_IDS:
                r = traci.lanearea.getSubscriptionResults(det)
                if r:
                    jl  = r.get(traci.constants.JAM_LENGTH_METERS, 0)
                    tp += r.get(traci.constants.VAR_INTERVAL_NUMBER, 0)
                    jam += jl
                    if det in NORTH_IDS:
                        north_jam += jl
                    else:
                        south_jam += jl
            jam_total       += jam / DETECTOR_COUNT
            north_jam_total += north_jam / len(NORTH_IDS)
            south_jam_total += south_jam / len(SOUTH_IDS)
            tp_total        += tp

        traci.simulationStep()

    sim_s = step * STEP_LENGTH
    traci.close()

    mean_queue    = jam_total       / obs if obs else 0.0
    mean_n_queue  = north_jam_total / obs if obs else 0.0
    mean_s_queue  = south_jam_total / obs if obs else 0.0
    throughput_hr = tp_total / (sim_s / 3600.0) if sim_s > 0 else 0.0
    mean_veh_tl, mean_ped_tl = parse_tripinfo(test["trips_out"])

    print(f"  Queue:{mean_queue:.2f}m | N:{mean_n_queue:.2f} S:{mean_s_queue:.2f} | "
          f"VehTL:{mean_veh_tl:.2f}s | PedTL:{mean_ped_tl:.2f}s | TP:{throughput_hr:.1f}veh/hr")
    return mean_veh_tl, mean_ped_tl, mean_queue, throughput_hr, mean_n_queue, mean_s_queue

# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

SCEN_CLR = {"normal": "D6E4F7", "slow": "FFF2CC", "jam": "FCE4D6"}
HDR_CLR  = "2E4057"
thin = Side(style="thin", color="AAAAAA")
BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)
COLS    = ["A", "B", "C", "D", "E", "F", "G"]
WIDTHS  = [18, 22, 22, 22, 20, 20, 20]
HEADERS = ["Run",
           "Mean Veh. Time Loss (s)", "Mean Ped. Time Loss (s)",
           "Mean Queue Length (m)",   "Throughput (veh/hr)",
           "Mean North Queue (m)",    "Mean South Queue (m)"]


def _hf(): return Font(bold=True, color="FFFFFF", name="Arial", size=11)
def _bf(): return Font(bold=True, name="Arial", size=10)
def _rf(): return Font(name="Arial", size=10)
def _c():  return Alignment(horizontal="center", vertical="center")
def _l():  return Alignment(horizontal="left",   vertical="center")


def build_xlsx(results, title, subtitle):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    for letter, w in zip(COLS, WIDTHS):
        ws.column_dimensions[letter].width = w
    last = COLS[-1]

    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, name="Arial", size=13)
    ws["A1"].alignment = _c(); ws.row_dimensions[1].height = 24

    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="555555")
    ws["A2"].alignment = _c(); ws.row_dimensions[2].height = 16

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _hf(); c.fill = PatternFill("solid", fgColor=HDR_CLR)
        c.alignment = _c(); c.border = BDR
    ws.row_dimensions[3].height = 18

    row = 4
    s_start, s_end, cur = {}, {}, None

    for res in results:
        scen = res["scenario"]
        if scen != cur:
            if cur: s_end[cur] = row - 1
            cur = scen; s_start[scen] = row
            gf = PatternFill("solid", fgColor=SCEN_CLR[scen])
            ws.merge_cells(f"A{row}:{last}{row}")
            ws[f"A{row}"] = f"— {scen.upper()} TRAFFIC SCENARIOS —"
            ws[f"A{row}"].font = Font(bold=True, name="Arial", size=10, color="333333")
            ws[f"A{row}"].fill = gf; ws[f"A{row}"].alignment = _c()
            ws[f"A{row}"].border = BDR; ws.row_dimensions[row].height = 15; row += 1

        sf = PatternFill("solid", fgColor=SCEN_CLR[scen])
        vals = [res["label"], res["mean_veh_tl"], res["mean_ped_tl"],
                res["queue"], res["throughput"], res["north_queue"], res["south_queue"]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = sf; c.border = BDR
            c.alignment = _l() if col == 1 else _c()
            c.font = _rf()
            if col > 1: c.number_format = "0.00"
        ws.row_dimensions[row].height = 14; row += 1

    if cur: s_end[cur] = row - 1

    row += 1
    ws.merge_cells(f"A{row}:{last}{row}")
    ws[f"A{row}"] = "SCENARIO SUMMARIES"
    ws[f"A{row}"].font = Font(bold=True, name="Arial", size=11, color=HDR_CLR)
    ws[f"A{row}"].alignment = _c(); ws.row_dimensions[row].height = 18; row += 1

    for col, h in enumerate(["Scenario"] + HEADERS[1:], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _hf(); c.fill = PatternFill("solid", fgColor=HDR_CLR)
        c.alignment = _c(); c.border = BDR
    ws.row_dimensions[row].height = 16; row += 1

    for scen in ["normal", "slow", "jam"]:
        s, e = s_start.get(scen), s_end.get(scen)
        if not s or not e: continue
        fill = PatternFill("solid", fgColor=SCEN_CLR[scen])
        c = ws.cell(row=row, column=1, value=f"{scen.capitalize()} Traffic")
        c.font = _bf(); c.fill = fill; c.alignment = _l(); c.border = BDR
        for col, letter in enumerate(["B", "C", "D", "E", "F", "G"], 2):
            c = ws.cell(row=row, column=col, value=f"=AVERAGE({letter}{s+1}:{letter}{e})")
            c.font = _bf(); c.fill = fill; c.alignment = _c()
            c.border = BDR; c.number_format = "0.00"
        ws.row_dimensions[row].height = 14; row += 1

    row += 1
    of = PatternFill("solid", fgColor="D5E8D4")
    ws.merge_cells(f"A{row}:{last}{row}")
    ws[f"A{row}"] = "OVERALL MEAN (all 30 runs)"
    ws[f"A{row}"].font = Font(bold=True, name="Arial", size=10, color="1B5E20")
    ws[f"A{row}"].fill = of; ws[f"A{row}"].alignment = _c()
    ws[f"A{row}"].border = BDR; ws.row_dimensions[row].height = 16; row += 1

    all_rows = []
    for scen in ["normal", "slow", "jam"]:
        s, e = s_start.get(scen), s_end.get(scen)
        if s and e: all_rows.extend(range(s + 1, e + 1))

    for col, letter in enumerate(COLS, 1):
        c = ws.cell(row=row, column=col)
        c.fill = of; c.border = BDR; c.alignment = _c()
        if col == 1:
            c.value = "All Scenarios"; c.font = _bf()
        else:
            refs = ",".join(f"{letter}{r}" for r in all_rows)
            c.value = f"=AVERAGE({refs})"
            c.font = _bf(); c.number_format = "0.00"
    ws.row_dimensions[row].height = 14
    ws.freeze_panes = "A4"
    return wb

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    results = []
    for i, test in enumerate(TEST_PLAN, 1):
        print(f"\n[{i}/{len(TEST_PLAN)}] {test['label']}")
        if not os.path.isfile(test["rou_file"]):
            print(f"  SKIP — not found: {test['rou_file']}")
            results.append({"label": test["label"], "scenario": test["scenario"],
                             "mean_veh_tl": None, "mean_ped_tl": None, "queue": None,
                             "throughput": None, "north_queue": None, "south_queue": None})
            continue
        try:
            vt, pt, ql, tp, nq, sq = run_simulation(test)
            results.append({"label": test["label"], "scenario": test["scenario"],
                             "mean_veh_tl": round(vt, 4), "mean_ped_tl": round(pt, 4),
                             "queue": round(ql, 4), "throughput": round(tp, 4),
                             "north_queue": round(nq, 4), "south_queue": round(sq, 4)})
        except Exception as e:
            print(f"  ERROR: {e}")
            results.append({"label": test["label"], "scenario": test["scenario"],
                             "mean_veh_tl": None, "mean_ped_tl": None, "queue": None,
                             "throughput": None, "north_queue": None, "south_queue": None})

    wb = build_xlsx(results,
                    "Balibago Signalized Pedestrian — DDPG Batch Test Results",
                    "30 Runs  |  10 Normal  •  10 Slow  •  10 Jam  |  DDPG trained model (test mode)")
    wb.save(XLSX_OUT)
    print(f"\n{'='*65}\n  Saved: {XLSX_OUT}\n{'='*65}")


if __name__ == "__main__":
    main()
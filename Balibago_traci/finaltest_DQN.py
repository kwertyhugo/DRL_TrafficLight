"""
finaltest_SP_DQN_integrated.py

Integrates the WORKING inline-loop pattern (from BP DQN) into the
Signalized-Pedestrian DQN batch test.

Root cause fixed:
  The original SP DQN used apply_north_phase() / apply_south_phase()
  helpers that advanced the phase counter *inside* the helper, then
  applied the action to the NEW phase rather than the current one.
  This is the same phase-offset bug fixed in the A2C scripts.
  Fix: phase transitions are done inline in the main loop, identical
  to the working BP DQN and the TestSignalized* single-run scripts.

SP-specific parameters preserved exactly:
  - signalizedPed.sumocfg
  - MAX_STEPS = 567000
  - North: 8-phase (%8), state=12 (8q + 4ph, no ped), norm /1000
    base = {0:45, 2:130, 4:30, 6:90}
  - South: 8-phase (%8), state=9  (5q + 4ph, no ped), norm /1000
    base = {0:25, 2:30,  4:40, 6:45}
  - Models: Balibago_traci/models_DQN/North_DQNAgent.keras
            Balibago_traci/models_DQN/South_DQNAgent.keras
  - verify_and_load() weight-check kept from original
"""

import os
import sys
import csv
import xml.etree.ElementTree as ET
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

if 'SUMO_HOME' in os.environ:
    sys.path.append(os.path.join(os.environ['SUMO_HOME'], 'tools'))
else:
    sys.exit("Please declare environment variable 'SUMO_HOME'")

import traci
from keras.utils import to_categorical
from models.DQN import DQNAgent as dqn

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
SUMOCFG          = 'Balibago_traci/signalizedPed.sumocfg'
DEMAND_DIR       = 'Balibago_traci/demand_test'
OUTPUT_DIR       = 'Balibago_traci/batch_results/SP_DQN'
XLSX_OUT         = 'Balibago_traci/batch_results/SP_DQN_Results.xlsx'
DEBUG_LOG        = 'Balibago_traci/batch_results/SP_DQN_debug.txt'
NORTH_MODEL_PATH = 'Balibago_traci/models_DQN/North_DQNAgent.keras'
SOUTH_MODEL_PATH = 'Balibago_traci/models_DQN/South_DQNAgent.keras'
STEP_LENGTH      = 0.1
MAX_STEPS        = 567000
METRIC_STEPS     = int(60 / STEP_LENGTH)

DETECTOR_IDS   = [f"e2_{i}" for i in range(13)]
DETECTOR_COUNT = 13
NORTH_IDS      = [f"e2_{i}" for i in range(8)]
SOUTH_IDS      = [f"e2_{i}" for i in range(8, 13)]

ACTION_SPACE   = (-25, -20, -15, -10, -5, 0, 5, 10, 15, 20, 25)  # index 5 = neutral (0)

NORTH_TL     = "4902876117"
NORTH_PHASES = 8
NORTH_BASE   = {0: 45, 2: 130, 4: 30, 6: 90}

SOUTH_TL     = "12188714"
SOUTH_PHASES = 8
SOUTH_BASE   = {0: 25, 2: 30, 4: 40, 6: 45}

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(os.path.dirname(XLSX_OUT), exist_ok=True)

# ---------------------------------------------------------------------------
# Load agents  (epsilon=0.0 → greedy, no exploration during test)
# ---------------------------------------------------------------------------
NorthAgent = dqn(state_size=12, action_size=11, memory_size=2000, gamma=0.95,
                 epsilon=0.0, epsilon_decay_rate=0.995, epsilon_min=0.01,
                 learning_rate=0.00005, target_update_freq=500,
                 name='North_DQNAgent', area='Balibago')

SouthAgent = dqn(state_size=9,  action_size=11, memory_size=2000, gamma=0.95,
                 epsilon=0.0, epsilon_decay_rate=0.995, epsilon_min=0.01,
                 learning_rate=0.00005, target_update_freq=500,
                 name='South_DQNAgent', area='Balibago')

# ---------------------------------------------------------------------------
# Model load + verification  (kept from original SP DQN)
# ---------------------------------------------------------------------------
print(f"\n{'='*70}")
print("  MODEL LOAD VERIFICATION — SP DQN")
print(f"{'='*70}")

def verify_and_load(agent, label, model_path):
    print(f"\n  [{label}]")
    print(f"    Expected path : {os.path.abspath(model_path)}")
    print(f"    State size    : {agent.state_size}")

    if not os.path.exists(model_path):
        print(f"    STATUS        : *** FILE NOT FOUND — weights will NOT be loaded! ***")
        return False

    print(f"    STATUS        : File found  ({os.path.getsize(model_path):,} bytes)")

    dummy    = np.zeros((1, agent.state_size), dtype=np.float32)
    q_before = agent.model.predict(dummy, verbose=0).copy()
    agent.model.load_weights(model_path)
    q_after  = agent.model.predict(dummy, verbose=0).copy()

    weights_changed = not np.allclose(q_before, q_after, atol=1e-6)
    if weights_changed:
        print(f"    Weights loaded : YES — Q-values changed after load (good)")
    else:
        print(f"    Weights loaded : *** UNCHANGED — load() may have silently failed ***")
        print(f"    >>> Q-before: {q_before}")
        print(f"    >>> Q-after : {q_after}")

    best = int(np.argmax(q_after))
    print(f"    Greedy action  : {best}  (ACTION_SPACE[{best}] = {ACTION_SPACE[best]:+d} s)")
    return weights_changed

north_ok = verify_and_load(NorthAgent, "North_DQNAgent  (state=12: 8q + 4ph, no ped)", NORTH_MODEL_PATH)
south_ok = verify_and_load(SouthAgent, "South_DQNAgent  (state=9:  5q + 4ph, no ped)", SOUTH_MODEL_PATH)

print(f"\n  Summary:")
print(f"    North model loaded properly : {'YES' if north_ok else '*** NO — RANDOM WEIGHTS ***'}")
print(f"    South model loaded properly : {'YES' if south_ok else '*** NO — RANDOM WEIGHTS ***'}")
if not north_ok or not south_ok:
    print(f"\n  *** WARNING: One or both agents did not load trained weights. ***")
    print(f"  *** Results will NOT reflect DQN performance.               ***")
print(f"\n{'='*70}\n  Verification complete.\n{'='*70}")

# ---------------------------------------------------------------------------
# Test plan  (10 normal + 10 slow + 10 jam = 30 runs)
# ---------------------------------------------------------------------------
TEST_PLAN = []
for scenario in ["normal", "slow", "jam"]:
    for i in range(1, 11):
        TEST_PLAN.append({
            "label":       f"{scenario.capitalize()} {i}",
            "scenario":    scenario,
            "rou_file":    os.path.join(DEMAND_DIR, f"flows_{scenario}_traffic_{i:02d}.rou.xml"),
            "trips_out":   os.path.join(OUTPUT_DIR,  f"SP_DQN_trips_{scenario}_{i:02d}.xml"),
            "stats_out":   os.path.join(OUTPUT_DIR,  f"SP_DQN_stats_{scenario}_{i:02d}.xml"),
            "metrics_csv": os.path.join(OUTPUT_DIR,  f"SP_DQN_metrics_{scenario}_{i:02d}.csv"),
        })

# ---------------------------------------------------------------------------
# Subscriptions  (no pedestrian subscriptions — SP DQN state has no ped term)
# ---------------------------------------------------------------------------

def _subscribe_all_detectors():
    ctx_vars = [traci.constants.VAR_TYPE, traci.constants.VAR_WAITING_TIME]
    met_vars = [traci.constants.JAM_LENGTH_METERS, traci.constants.VAR_INTERVAL_NUMBER]
    for det in DETECTOR_IDS:
        traci.lanearea.subscribeContext(det, traci.constants.CMD_GET_VEHICLE_VARIABLE, 3, ctx_vars)
        traci.lanearea.subscribe(det, met_vars)

# ---------------------------------------------------------------------------
# State helpers — exact match to trained SP DQN model
# North: 8q + 4ph = 12, norm /1000, num_classes=4
# South: 5q + 4ph =  9, norm /1000, num_classes=4
# ---------------------------------------------------------------------------

def _weighted_waits(det_id):
    data = traci.lanearea.getContextSubscriptionResults(det_id)
    if not data:
        return 0
    weights = {"car": 1.0, "jeep": 1.5, "bus": 2.2,
               "truck": 2.5, "motorcycle": 0.3, "tricycle": 0.5}
    sumWait = 0
    for d in data.values():
        v_type   = d.get(traci.constants.VAR_TYPE, "car")
        waitTime = d.get(traci.constants.VAR_WAITING_TIME, 0)
        sumWait += waitTime * weights.get(v_type, 1.0)
    return sumWait

def _northIntersection_queue():
    return [_weighted_waits(f"e2_{i}") for i in range(8)]

def _southIntersection_queue():
    return [_weighted_waits(f"e2_{i}") for i in range(8, 13)]

# ---------------------------------------------------------------------------
# Metrics CSV saver
# ---------------------------------------------------------------------------

def save_metrics_csv(filename, metrics_list):
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    with open(filename, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Time_min', 'Avg_Jam_Length_m', 'Throughput_veh_per_min',
                         'North_Queue', 'South_Queue', 'Total_Queue',
                         'North_Jam_Length_m', 'South_Jam_Length_m'])
        writer.writerows(metrics_list)

# ---------------------------------------------------------------------------
# Tripinfo parser
# ---------------------------------------------------------------------------

def parse_tripinfo(path):
    try:
        root = ET.parse(path).getroot()
    except (FileNotFoundError, ET.ParseError):
        return 0.0, 0.0
    veh, ped = [], []
    for tag in root:
        tl = tag.get("timeLoss")
        if tl is None:
            continue
        (veh if tag.tag == "tripinfo" else ped).append(float(tl))
    return (sum(veh) / len(veh) if veh else 0.0,
            sum(ped) / len(ped) if ped else 0.0)

# ---------------------------------------------------------------------------
# Single simulation run — inline loop (fixes phase-offset bug)
# ---------------------------------------------------------------------------

def run_simulation(test):
    sumo_cmd = [
        'sumo', '-c', SUMOCFG,
        '--route-files',        test["rou_file"],
        '--step-length',        str(STEP_LENGTH),
        '--delay',              '0',
        '--lateral-resolution', '0.1',
        '--statistic-output',   test["stats_out"],
        '--tripinfo-output',    test["trips_out"],
        '--no-warnings',        'true',
    ]
    print(f"\n{'='*70}\n  Running: {test['label']}\n{'='*70}")
    traci.start(sumo_cmd)
    _subscribe_all_detectors()

    # ── Phase state ───────────────────────────────────────────────────────────
    northCurrentPhase         = 0
    northCurrentPhaseDuration = 30.0
    southCurrentPhase         = 0
    southCurrentPhaseDuration = 30.0

    step_counter             = 0
    metric_observation_count = 0
    throughput_total         = 0
    jam_length_total         = 0.0
    total_queue_north        = 0.0
    total_queue_south        = 0.0
    north_jam_length_total   = 0.0
    south_jam_length_total   = 0.0

    metrics_timeline = []
    north_actions    = []
    south_actions    = []

    while traci.simulation.getMinExpectedNumber() > 0 and step_counter < MAX_STEPS:
        step_counter              += 1
        northCurrentPhaseDuration -= STEP_LENGTH
        southCurrentPhaseDuration -= STEP_LENGTH

        north_decision_needed = (northCurrentPhaseDuration <= 0) and (northCurrentPhase % 2 == 0)
        south_decision_needed = (southCurrentPhaseDuration <= 0) and (southCurrentPhase % 2 == 0)

        next_action_N_idx = None
        next_action_S_idx = None

        # ── North agent decision ─────────────────────────────────────────────
        # State: 8q + 4 phase OH = 12, norm /1000, num_classes=4
        if north_decision_needed:
            queue        = np.array(_northIntersection_queue())
            norm_q_north = queue / 1000.0
            n_phase_oh   = to_categorical(northCurrentPhase // 2, num_classes=4).flatten()
            obs_north    = np.concatenate([norm_q_north, n_phase_oh]).astype(np.float32)
            next_action_N_idx = NorthAgent.act(obs_north)
            north_actions.append(next_action_N_idx)
        elif northCurrentPhaseDuration <= 0:
            next_action_N_idx = 5  # neutral — yellow transition

        # ── South agent decision ─────────────────────────────────────────────
        # State: 5q + 4 phase OH = 9, norm /1000, num_classes=4
        if south_decision_needed:
            queue        = np.array(_southIntersection_queue())
            norm_q_south = queue / 1000.0
            s_phase_oh   = to_categorical(southCurrentPhase // 2, num_classes=4).flatten()
            obs_south    = np.concatenate([norm_q_south, s_phase_oh]).astype(np.float32)
            next_action_S_idx = SouthAgent.act(obs_south)
            south_actions.append(next_action_S_idx)
        elif southCurrentPhaseDuration <= 0:
            next_action_S_idx = 5  # neutral — yellow transition

        # ── Apply north phase transition (8-phase cycle) ─────────────────────
        if northCurrentPhaseDuration <= 0:
            northCurrentPhase = (northCurrentPhase + 1) % NORTH_PHASES
            traci.trafficlight.setPhase(NORTH_TL, northCurrentPhase)

            if northCurrentPhase % 2 == 1:
                northCurrentPhaseDuration = 5.0          # yellow
            else:
                duration_adj = ACTION_SPACE[next_action_N_idx]
                base = NORTH_BASE.get(northCurrentPhase, 30)
                northCurrentPhaseDuration = float(max(5, min(180, base + duration_adj)))

            traci.trafficlight.setPhaseDuration(NORTH_TL, northCurrentPhaseDuration)

        # ── Apply south phase transition (8-phase cycle) ─────────────────────
        if southCurrentPhaseDuration <= 0:
            southCurrentPhase = (southCurrentPhase + 1) % SOUTH_PHASES
            traci.trafficlight.setPhase(SOUTH_TL, southCurrentPhase)

            if southCurrentPhase % 2 == 1:
                southCurrentPhaseDuration = 5.0          # yellow
            else:
                duration_adj = ACTION_SPACE[next_action_S_idx]
                base = SOUTH_BASE.get(southCurrentPhase, 30)
                southCurrentPhaseDuration = float(max(5, min(180, base + duration_adj)))

            traci.trafficlight.setPhaseDuration(SOUTH_TL, southCurrentPhaseDuration)

        # ── Collect per-minute metrics ────────────────────────────────────────
        if step_counter % METRIC_STEPS == 0:
            jam_length       = 0.0
            throughput       = 0
            north_jam_length = 0.0
            south_jam_length = 0.0
            metric_observation_count += 1

            for i, det_id in enumerate(DETECTOR_IDS):
                det_stats = traci.lanearea.getSubscriptionResults(det_id)
                if not det_stats:
                    continue
                det_jam        = det_stats.get(traci.constants.JAM_LENGTH_METERS, 0)
                det_throughput = det_stats.get(traci.constants.VAR_INTERVAL_NUMBER, 0)
                jam_length    += det_jam
                throughput    += det_throughput
                if i < 8:
                    north_jam_length += det_jam
                else:
                    south_jam_length += det_jam

            jam_length   /= DETECTOR_COUNT
            jam_length_total          += jam_length
            throughput_total          += throughput

            north_jam_length /= len(NORTH_IDS)
            south_jam_length /= len(SOUTH_IDS)
            north_jam_length_total    += north_jam_length
            south_jam_length_total    += south_jam_length

            north_queue = sum(_northIntersection_queue())
            south_queue = sum(_southIntersection_queue())
            total_queue_north         += north_queue
            total_queue_south         += south_queue

            time_min = step_counter * STEP_LENGTH / 60.0
            metrics_timeline.append([
                f"{time_min:.1f}",
                f"{jam_length:.2f}",
                f"{throughput:.2f}",
                f"{north_queue:.2f}",
                f"{south_queue:.2f}",
                f"{north_queue + south_queue:.2f}",
                f"{north_jam_length:.2f}",
                f"{south_jam_length:.2f}",
            ])

        traci.simulationStep()

    sim_s = step_counter * STEP_LENGTH
    traci.close()

    # ── Compute run averages ──────────────────────────────────────────────────
    n = metric_observation_count if metric_observation_count > 0 else 1
    avg_jam         = jam_length_total       / n
    avg_throughput  = throughput_total       / n
    avg_queue_north = total_queue_north      / n
    avg_queue_south = total_queue_south      / n
    avg_north_jam   = north_jam_length_total / n
    avg_south_jam   = south_jam_length_total / n

    throughput_hr   = (throughput_total / (sim_s / 3600.0)) if sim_s > 0 else 0.0
    mean_veh_tl, mean_ped_tl = parse_tripinfo(test["trips_out"])

    # ── Print + debug log ─────────────────────────────────────────────────────
    n_action_dist = Counter(north_actions)
    s_action_dist = Counter(south_actions)
    debug_msg = (
        f"\n  Results for {test['label']}:\n"
        f"    Average Jam Length (Overall)    : {avg_jam:.2f} m\n"
        f"    Average Jam Length (North)      : {avg_north_jam:.2f} m\n"
        f"    Average Jam Length (South)      : {avg_south_jam:.2f} m\n"
        f"    Average Throughput (per min)    : {avg_throughput:.2f} veh/min\n"
        f"    Throughput (per hr)             : {throughput_hr:.1f} veh/hr\n"
        f"    Average North Queue             : {avg_queue_north:.2f}\n"
        f"    Average South Queue             : {avg_queue_south:.2f}\n"
        f"    Average Total Queue             : {avg_queue_north + avg_queue_south:.2f}\n"
        f"    Mean Veh. Time Loss             : {mean_veh_tl:.2f} s\n"
        f"    Mean Ped. Time Loss             : {mean_ped_tl:.2f} s\n"
        f"    North decisions: {len(north_actions)}, unique actions: {len(set(north_actions))}/11, "
        f"distribution: {dict(n_action_dist)}\n"
        f"    South decisions: {len(south_actions)}, unique actions: {len(set(south_actions))}/11, "
        f"distribution: {dict(s_action_dist)}"
    )
    print(debug_msg)

    with open(DEBUG_LOG, 'a') as f:
        f.write(f"\n{'='*70}\n{test['label']}\n{debug_msg}\n")

    save_metrics_csv(test["metrics_csv"], metrics_timeline)
    print(f"    Saved metrics CSV to {test['metrics_csv']}")

    return mean_veh_tl, mean_ped_tl, avg_jam, throughput_hr, avg_queue_north, avg_queue_south

# ---------------------------------------------------------------------------
# Excel builder  (unchanged from original SP DQN)
# ---------------------------------------------------------------------------

SCEN_CLR = {"normal": "D6E4F7", "slow": "FFF2CC", "jam": "FCE4D6"}
HDR_CLR  = "2E4057"
thin = Side(style="thin", color="AAAAAA")
BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)
COLS    = ["A", "B", "C", "D", "E", "F", "G"]
WIDTHS  = [18, 22, 22, 22, 20, 20, 20]
HEADERS = ["Run",
           "Mean Veh. Time Loss (s)", "Mean Ped. Time Loss (s)",
           "Mean Queue Length (m)",   "Throughput (veh/hr)",
           "Mean North Queue (m)",    "Mean South Queue (m)"]

def _hf(): return Font(bold=True, color="FFFFFF", name="Arial", size=11)
def _bf(): return Font(bold=True, name="Arial", size=10)
def _rf(): return Font(name="Arial", size=10)
def _c():  return Alignment(horizontal="center", vertical="center")
def _l():  return Alignment(horizontal="left",   vertical="center")

def build_xlsx(results, title, subtitle):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    for letter, w in zip(COLS, WIDTHS):
        ws.column_dimensions[letter].width = w
    last = COLS[-1]

    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, name="Arial", size=13)
    ws["A1"].alignment = _c(); ws.row_dimensions[1].height = 24

    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="555555")
    ws["A2"].alignment = _c(); ws.row_dimensions[2].height = 16

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _hf(); c.fill = PatternFill("solid", fgColor=HDR_CLR)
        c.alignment = _c(); c.border = BDR
    ws.row_dimensions[3].height = 18

    row = 4
    s_start, s_end, cur = {}, {}, None

    for res in results:
        scen = res["scenario"]
        if scen != cur:
            if cur: s_end[cur] = row - 1
            cur = scen; s_start[scen] = row
            gf = PatternFill("solid", fgColor=SCEN_CLR[scen])
            ws.merge_cells(f"A{row}:{last}{row}")
            ws[f"A{row}"] = f"— {scen.upper()} TRAFFIC SCENARIOS —"
            ws[f"A{row}"].font = Font(bold=True, name="Arial", size=10, color="333333")
            ws[f"A{row}"].fill = gf; ws[f"A{row}"].alignment = _c()
            ws[f"A{row}"].border = BDR; ws.row_dimensions[row].height = 15; row += 1

        sf = PatternFill("solid", fgColor=SCEN_CLR[scen])
        vals = [res["label"], res["mean_veh_tl"], res["mean_ped_tl"],
                res["queue"], res["throughput"], res["north_queue"], res["south_queue"]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = sf; c.border = BDR
            c.alignment = _l() if col == 1 else _c()
            c.font = _rf()
            if col > 1: c.number_format = "0.00"
        ws.row_dimensions[row].height = 14; row += 1

    if cur: s_end[cur] = row - 1

    row += 1
    ws.merge_cells(f"A{row}:{last}{row}")
    ws[f"A{row}"] = "SCENARIO SUMMARIES"
    ws[f"A{row}"].font = Font(bold=True, name="Arial", size=11, color=HDR_CLR)
    ws[f"A{row}"].alignment = _c(); ws.row_dimensions[row].height = 18; row += 1

    for col, h in enumerate(["Scenario"] + HEADERS[1:], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _hf(); c.fill = PatternFill("solid", fgColor=HDR_CLR)
        c.alignment = _c(); c.border = BDR
    ws.row_dimensions[row].height = 16; row += 1

    for scen in ["normal", "slow", "jam"]:
        s, e = s_start.get(scen), s_end.get(scen)
        if not s or not e: continue
        fill = PatternFill("solid", fgColor=SCEN_CLR[scen])
        c = ws.cell(row=row, column=1, value=f"{scen.capitalize()} Traffic")
        c.font = _bf(); c.fill = fill; c.alignment = _l(); c.border = BDR
        for col, letter in enumerate(["B", "C", "D", "E", "F", "G"], 2):
            c = ws.cell(row=row, column=col, value=f"=AVERAGE({letter}{s+1}:{letter}{e})")
            c.font = _bf(); c.fill = fill; c.alignment = _c()
            c.border = BDR; c.number_format = "0.00"
        ws.row_dimensions[row].height = 14; row += 1

    row += 1
    of = PatternFill("solid", fgColor="D5E8D4")
    ws.merge_cells(f"A{row}:{last}{row}")
    ws[f"A{row}"] = "OVERALL MEAN (all 30 runs)"
    ws[f"A{row}"].font = Font(bold=True, name="Arial", size=10, color="1B5E20")
    ws[f"A{row}"].fill = of; ws[f"A{row}"].alignment = _c()
    ws[f"A{row}"].border = BDR; ws.row_dimensions[row].height = 16; row += 1

    all_rows = []
    for scen in ["normal", "slow", "jam"]:
        s, e = s_start.get(scen), s_end.get(scen)
        if s and e: all_rows.extend(range(s + 1, e + 1))

    for col, letter in enumerate(COLS, 1):
        c = ws.cell(row=row, column=col)
        c.fill = of; c.border = BDR; c.alignment = _c()
        if col == 1:
            c.value = "All Scenarios"; c.font = _bf()
        else:
            refs = ",".join(f"{letter}{r}" for r in all_rows)
            c.value = f"=AVERAGE({refs})"
            c.font = _bf(); c.number_format = "0.00"
    ws.row_dimensions[row].height = 14
    ws.freeze_panes = "A4"
    return wb

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    with open(DEBUG_LOG, 'w') as f:
        f.write("SP DQN Batch Test Debug Log\n")
        f.write(f"Action Space: {ACTION_SPACE}\n")
        f.write(f"Neutral action index: 5 (adjustment = 0)\n")
        f.write(f"North: 8-phase, state=12 (8q + 4 OH), norm /1000\n")
        f.write(f"South: 8-phase, state=9  (5q + 4 OH), norm /1000\n")
        f.write("="*70 + "\n")

    results = []
    for i, test in enumerate(TEST_PLAN, 1):
        print(f"\n[{i}/{len(TEST_PLAN)}] {test['label']}")
        if not os.path.isfile(test["rou_file"]):
            print(f"  SKIP — route file not found: {test['rou_file']}")
            results.append({"label": test["label"], "scenario": test["scenario"],
                             "mean_veh_tl": None, "mean_ped_tl": None,
                             "queue": None, "throughput": None,
                             "north_queue": None, "south_queue": None})
            continue
        try:
            vt, pt, ql, tp, nq, sq = run_simulation(test)
            results.append({"label": test["label"], "scenario": test["scenario"],
                             "mean_veh_tl": round(vt, 4), "mean_ped_tl": round(pt, 4),
                             "queue": round(ql, 4), "throughput": round(tp, 4),
                             "north_queue": round(nq, 4), "south_queue": round(sq, 4)})
        except Exception as e:
            print(f"  ERROR: {e}")
            import traceback
            traceback.print_exc()
            try:
                traci.close()
            except:
                pass
            results.append({"label": test["label"], "scenario": test["scenario"],
                             "mean_veh_tl": None, "mean_ped_tl": None,
                             "queue": None, "throughput": None,
                             "north_queue": None, "south_queue": None})

    wb = build_xlsx(
        results,
        "Balibago Signalized Pedestrian — DQN Batch Test Results",
        "30 Runs  |  10 Normal  •  10 Slow  •  10 Jam  |  DQN trained model (integrated loop)"
    )
    wb.save(XLSX_OUT)
    print(f"\n{'='*70}\n  Saved Excel: {XLSX_OUT}\n  Debug log:   {DEBUG_LOG}\n{'='*70}")

if __name__ == "__main__":
    main()
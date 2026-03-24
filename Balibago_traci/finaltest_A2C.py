"""
finaltest_A2C_integrated.py

Integrates the WORKING simulation loop from TestSignalizedA2CSlow into
the 30-run batch framework of finaltest_A2C.

Key changes from finaltest_A2C:
- Simulation loop mirrors TestSignalizedA2CSlow exactly (phase timing,
  state construction, decision guards, apply logic all in-loop — NOT
  broken into helper apply_*_phase functions that introduced the bug)
- Metrics now also include per-minute timeline CSV per run (same as
  TestSignalizedA2CSlow), saved alongside the XML outputs
- Excel output and batch structure kept from finaltest_A2C unchanged
"""

import os
import sys
import csv
import xml.etree.ElementTree as ET
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

if 'SUMO_HOME' in os.environ:
    sys.path.append(os.path.join(os.environ['SUMO_HOME'], 'tools'))
else:
    sys.exit("Please declare environment variable 'SUMO_HOME'")

import traci
from keras.utils import to_categorical
from keras.models import load_model

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
SUMOCFG          = 'Balibago_traci/signalizedPed.sumocfg'
DEMAND_DIR       = 'Balibago_traci/demand_test'
OUTPUT_DIR       = 'Balibago_traci/batch_results/SP_A2C'
XLSX_OUT         = 'Balibago_traci/batch_results/SP_A2C_Results.xlsx'
DEBUG_LOG        = 'Balibago_traci/batch_results/SP_A2C_debug.txt'
NORTH_MODEL_PATH = './Balibago_traci/models_A2C/North_A2CAgent.keras'
SOUTH_MODEL_PATH = './Balibago_traci/models_A2C/South_A2CAgent.keras'
STEP_LENGTH      = 0.1
MAX_STEPS        = 576000
METRIC_STEPS     = int(60 / STEP_LENGTH)   # 1-minute observation intervals

DETECTOR_IDS   = [f"e2_{i}" for i in range(13)]
DETECTOR_COUNT = 13
NORTH_IDS      = [f"e2_{i}" for i in range(8)]
SOUTH_IDS      = [f"e2_{i}" for i in range(8, 13)]

ACTION_SPACE   = (-25, -20, -15, -10, -5, 0, 5, 10, 15, 20, 25)  # index 5 = neutral (0)

NORTH_TL     = "4902876117"
NORTH_BASE   = {0: 45, 2: 130, 4: 30, 6: 90}

SOUTH_TL     = "12188714"
SOUTH_BASE   = {0: 25, 2: 30, 4: 40, 6: 45}

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(os.path.dirname(XLSX_OUT), exist_ok=True)

# ---------------------------------------------------------------------------
# Load models
# ---------------------------------------------------------------------------
print(f"\n{'='*70}")
print("Loading Trained A2C Models")
print('='*70)
for path, label in [(NORTH_MODEL_PATH, "North"), (SOUTH_MODEL_PATH, "South")]:
    if not os.path.exists(path):
        sys.exit(f"ERROR: {label} model not found at {path}")
north_model = load_model(NORTH_MODEL_PATH)
south_model = load_model(SOUTH_MODEL_PATH)
print(f"  ✓ Loaded North Agent from {NORTH_MODEL_PATH}")
print(f"  ✓ Loaded South Agent from {SOUTH_MODEL_PATH}")
print('='*70)

# ---------------------------------------------------------------------------
# Test plan  (10 normal + 10 slow + 10 jam = 30 runs)
# ---------------------------------------------------------------------------
TEST_PLAN = []
for scenario in ["normal", "slow", "jam"]:
    for i in range(1, 11):
        TEST_PLAN.append({
            "label":      f"{scenario.capitalize()} {i}",
            "scenario":   scenario,
            "rou_file":   os.path.join(DEMAND_DIR, f"flows_{scenario}_traffic_{i:02d}.rou.xml"),
            "trips_out":  os.path.join(OUTPUT_DIR,  f"SP_A2C_trips_{scenario}_{i:02d}.xml"),
            "stats_out":  os.path.join(OUTPUT_DIR,  f"SP_A2C_stats_{scenario}_{i:02d}.xml"),
            "metrics_csv":os.path.join(OUTPUT_DIR,  f"SP_A2C_metrics_{scenario}_{i:02d}.csv"),
        })

# ---------------------------------------------------------------------------
# Subscriptions  (identical to TestSignalizedA2CSlow)
# ---------------------------------------------------------------------------

def _subscribe_all_detectors():
    ctx_vars = [traci.constants.VAR_TYPE, traci.constants.VAR_WAITING_TIME]
    met_vars = [traci.constants.JAM_LENGTH_METERS, traci.constants.VAR_INTERVAL_NUMBER]
    for det in DETECTOR_IDS:
        traci.lanearea.subscribeContext(det, traci.constants.CMD_GET_VEHICLE_VARIABLE, 3, ctx_vars)
        traci.lanearea.subscribe(det, met_vars)

def _junctionSubscription(junction_id):
    traci.junction.subscribeContext(
        junction_id,
        traci.constants.CMD_GET_PERSON_VARIABLE,
        10.0,
        [traci.constants.VAR_WAITING_TIME]
    )

# ---------------------------------------------------------------------------
# State helpers  (identical to TestSignalizedA2CSlow)
# ---------------------------------------------------------------------------

def _weighted_waits(det_id):
    data = traci.lanearea.getContextSubscriptionResults(det_id)
    if not data:
        return 0
    weights = {"car": 1.0, "jeep": 1.5, "bus": 2.2,
               "truck": 2.5, "motorcycle": 0.3, "tricycle": 0.5}
    sumWait = 0
    for d in data.values():
        v_type   = d.get(traci.constants.VAR_TYPE, "car")
        waitTime = d.get(traci.constants.VAR_WAITING_TIME, 0)
        sumWait += waitTime * weights.get(v_type, 1.0)
    return sumWait

def _northIntersection_queue():
    queues = [_weighted_waits(f"e2_{i}") for i in range(8)]
    pedestrian = 0
    junction_data = traci.junction.getContextSubscriptionResults(NORTH_TL)
    if junction_data:
        for d in junction_data.values():
            pedestrian += d.get(traci.constants.VAR_WAITING_TIME, 0)
    return queues + [pedestrian]

def _southIntersection_queue():
    queues = [_weighted_waits(f"e2_{i}") for i in range(8, 13)]
    pedestrian = 0
    junction_data = traci.junction.getContextSubscriptionResults(SOUTH_TL)
    if junction_data:
        for d in junction_data.values():
            pedestrian += d.get(traci.constants.VAR_WAITING_TIME, 0)
    return queues + [pedestrian]

def predict_action(model, state):
    """Greedy action from trained model — no exploration."""
    state_batch = np.expand_dims(state, axis=0)
    action_probs = model.predict(state_batch, verbose=0)[0]
    return int(np.argmax(action_probs))

# ---------------------------------------------------------------------------
# Metrics CSV saver
# ---------------------------------------------------------------------------

def save_metrics_csv(filename, metrics_list):
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    with open(filename, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Time_min', 'Avg_Jam_Length_m', 'Throughput_veh_per_min',
                         'North_Queue', 'South_Queue', 'Total_Queue',
                         'North_Jam_Length_m', 'South_Jam_Length_m'])
        writer.writerows(metrics_list)

# ---------------------------------------------------------------------------
# Tripinfo parser  (from finaltest_A2C)
# ---------------------------------------------------------------------------

def parse_tripinfo(path):
    try:
        root = ET.parse(path).getroot()
    except (FileNotFoundError, ET.ParseError):
        return 0.0, 0.0
    veh, ped = [], []
    for tag in root:
        tl = tag.get("timeLoss")
        if tl is None:
            continue
        (veh if tag.tag == "tripinfo" else ped).append(float(tl))
    return (sum(veh) / len(veh) if veh else 0.0,
            sum(ped) / len(ped) if ped else 0.0)

# ---------------------------------------------------------------------------
# Single simulation run
# Uses the WORKING loop structure from TestSignalizedA2CSlow
# ---------------------------------------------------------------------------

def run_simulation(test):
    sumo_cmd = [
        'sumo', '-c', SUMOCFG,
        '--route-files',        test["rou_file"],
        '--step-length',        str(STEP_LENGTH),
        '--delay',              '0',
        '--lateral-resolution', '0.1',
        '--statistic-output',   test["stats_out"],
        '--tripinfo-output',    test["trips_out"],
        '--no-warnings',        'true',
    ]
    print(f"\n{'='*70}\n  Running: {test['label']}\n{'='*70}")
    traci.start(sumo_cmd)
    _subscribe_all_detectors()
    _junctionSubscription(NORTH_TL)
    _junctionSubscription(SOUTH_TL)

    # ── Phase state (mirrors TestSignalizedA2CSlow exactly) ──────────────────
    northCurrentPhase         = 0
    northCurrentPhaseDuration = 30.0
    southCurrentPhase         = 0
    southCurrentPhaseDuration = 30.0

    step_counter             = 0
    metric_observation_count = 0
    throughput_total         = 0
    jam_length_total         = 0.0
    total_queue_north        = 0.0
    total_queue_south        = 0.0
    north_jam_length_total   = 0.0
    south_jam_length_total   = 0.0

    metrics_timeline = []

    # Debug counters
    north_actions  = []
    south_actions  = []

    while traci.simulation.getMinExpectedNumber() > 0 and step_counter < MAX_STEPS:
        step_counter              += 1
        northCurrentPhaseDuration -= STEP_LENGTH
        southCurrentPhaseDuration -= STEP_LENGTH

        north_decision_needed = (northCurrentPhaseDuration <= 0) and (northCurrentPhase % 2 == 0)
        south_decision_needed = (southCurrentPhaseDuration <= 0) and (southCurrentPhase % 2 == 0)

        next_action_N_idx = None
        next_action_S_idx = None

        # ── North agent decision ─────────────────────────────────────────────
        if north_decision_needed:
            queue        = np.array(_northIntersection_queue())
            norm_q_north = queue / 2000.0
            n_phase_oh   = to_categorical(northCurrentPhase // 2, num_classes=4).flatten()
            obs_north    = np.concatenate([norm_q_north, n_phase_oh]).astype(np.float32)
            next_action_N_idx = predict_action(north_model, obs_north)
            north_actions.append(next_action_N_idx)
        elif northCurrentPhaseDuration <= 0:
            next_action_N_idx = 5  # neutral — yellow transition

        # ── South agent decision ─────────────────────────────────────────────
        if south_decision_needed:
            queue        = np.array(_southIntersection_queue())
            norm_q_south = queue / 1000.0
            s_phase_oh   = to_categorical(southCurrentPhase // 2, num_classes=4).flatten()
            obs_south    = np.concatenate([norm_q_south, s_phase_oh]).astype(np.float32)
            next_action_S_idx = predict_action(south_model, obs_south)
            south_actions.append(next_action_S_idx)
        elif southCurrentPhaseDuration <= 0:
            next_action_S_idx = 5  # neutral — yellow transition

        # ── Apply north phase transition ─────────────────────────────────────
        if northCurrentPhaseDuration <= 0:
            northCurrentPhase = (northCurrentPhase + 1) % 8
            traci.trafficlight.setPhase(NORTH_TL, northCurrentPhase)

            if northCurrentPhase % 2 == 1:
                northCurrentPhaseDuration = 5.0          # yellow
            else:
                duration_adj = ACTION_SPACE[next_action_N_idx]
                base = NORTH_BASE.get(northCurrentPhase, 30)
                northCurrentPhaseDuration = float(max(5, min(180, base + duration_adj)))

            traci.trafficlight.setPhaseDuration(NORTH_TL, northCurrentPhaseDuration)

        # ── Apply south phase transition ─────────────────────────────────────
        if southCurrentPhaseDuration <= 0:
            southCurrentPhase = (southCurrentPhase + 1) % 8
            traci.trafficlight.setPhase(SOUTH_TL, southCurrentPhase)

            if southCurrentPhase % 2 == 1:
                southCurrentPhaseDuration = 5.0          # yellow
            else:
                duration_adj = ACTION_SPACE[next_action_S_idx]
                base = SOUTH_BASE.get(southCurrentPhase, 30)
                southCurrentPhaseDuration = float(max(5, min(180, base + duration_adj)))

            traci.trafficlight.setPhaseDuration(SOUTH_TL, southCurrentPhaseDuration)

        # ── Collect per-minute metrics ────────────────────────────────────────
        if step_counter % METRIC_STEPS == 0:
            jam_length       = 0.0
            throughput       = 0
            north_jam_length = 0.0
            south_jam_length = 0.0
            metric_observation_count += 1

            for i, det_id in enumerate(DETECTOR_IDS):
                det_stats = traci.lanearea.getSubscriptionResults(det_id)
                if not det_stats:
                    continue
                det_jam        = det_stats.get(traci.constants.JAM_LENGTH_METERS, 0)
                det_throughput = det_stats.get(traci.constants.VAR_INTERVAL_NUMBER, 0)
                jam_length    += det_jam
                throughput    += det_throughput
                if i < 8:
                    north_jam_length += det_jam
                else:
                    south_jam_length += det_jam

            jam_length   /= DETECTOR_COUNT
            jam_length_total          += jam_length
            throughput_total          += throughput

            north_jam_length /= len(NORTH_IDS)
            south_jam_length /= len(SOUTH_IDS)
            north_jam_length_total    += north_jam_length
            south_jam_length_total    += south_jam_length

            north_queue = sum(_northIntersection_queue())
            south_queue = sum(_southIntersection_queue())
            total_queue_north         += north_queue
            total_queue_south         += south_queue

            time_min = step_counter * STEP_LENGTH / 60.0
            metrics_timeline.append([
                f"{time_min:.1f}",
                f"{jam_length:.2f}",
                f"{throughput:.2f}",
                f"{north_queue:.2f}",
                f"{south_queue:.2f}",
                f"{north_queue + south_queue:.2f}",
                f"{north_jam_length:.2f}",
                f"{south_jam_length:.2f}",
            ])

        traci.simulationStep()

    sim_s = step_counter * STEP_LENGTH
    traci.close()

    # ── Compute run averages ──────────────────────────────────────────────────
    n = metric_observation_count if metric_observation_count > 0 else 1
    avg_jam         = jam_length_total          / n
    avg_throughput  = throughput_total          / n        # veh / min avg
    avg_queue_north = total_queue_north         / n
    avg_queue_south = total_queue_south         / n
    avg_north_jam   = north_jam_length_total    / n
    avg_south_jam   = south_jam_length_total    / n

    # Convert mean throughput to veh/hr for Excel (matches finaltest_A2C)
    throughput_hr = (throughput_total / (sim_s / 3600.0)) if sim_s > 0 else 0.0

    mean_veh_tl, mean_ped_tl = parse_tripinfo(test["trips_out"])

    # ── Print + debug log ─────────────────────────────────────────────────────
    n_action_dist = Counter(north_actions)
    s_action_dist = Counter(south_actions)
    debug_msg = (
        f"\n  Results for {test['label']}:\n"
        f"    Average Jam Length (Overall)    : {avg_jam:.2f} m\n"
        f"    Average Jam Length (North)      : {avg_north_jam:.2f} m\n"
        f"    Average Jam Length (South)      : {avg_south_jam:.2f} m\n"
        f"    Average Throughput (per min)    : {avg_throughput:.2f} veh/min\n"
        f"    Throughput (per hr)             : {throughput_hr:.1f} veh/hr\n"
        f"    Average North Queue             : {avg_queue_north:.2f}\n"
        f"    Average South Queue             : {avg_queue_south:.2f}\n"
        f"    Average Total Queue             : {avg_queue_north + avg_queue_south:.2f}\n"
        f"    Mean Veh. Time Loss             : {mean_veh_tl:.2f} s\n"
        f"    Mean Ped. Time Loss             : {mean_ped_tl:.2f} s\n"
        f"    North decisions: {len(north_actions)}, unique actions: {len(set(north_actions))}/11, "
        f"distribution: {dict(n_action_dist)}\n"
        f"    South decisions: {len(south_actions)}, unique actions: {len(set(south_actions))}/11, "
        f"distribution: {dict(s_action_dist)}"
    )
    print(debug_msg)

    with open(DEBUG_LOG, 'a') as f:
        f.write(f"\n{'='*70}\n{test['label']}\n{debug_msg}\n")

    # ── Save per-minute metrics CSV ───────────────────────────────────────────
    save_metrics_csv(test["metrics_csv"], metrics_timeline)
    print(f"    ✓ Saved metrics CSV to {test['metrics_csv']}")

    return mean_veh_tl, mean_ped_tl, avg_jam, throughput_hr, avg_queue_north, avg_queue_south

# ---------------------------------------------------------------------------
# Excel builder  (unchanged from finaltest_A2C)
# ---------------------------------------------------------------------------

SCEN_CLR = {"normal": "D6E4F7", "slow": "FFF2CC", "jam": "FCE4D6"}
HDR_CLR  = "2E4057"
thin = Side(style="thin", color="AAAAAA")
BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)
COLS    = ["A", "B", "C", "D", "E", "F", "G"]
WIDTHS  = [18, 22, 22, 22, 20, 20, 20]
HEADERS = ["Run",
           "Mean Veh. Time Loss (s)", "Mean Ped. Time Loss (s)",
           "Mean Queue Length (m)",   "Throughput (veh/hr)",
           "Mean North Queue (m)",    "Mean South Queue (m)"]

def _hf(): return Font(bold=True, color="FFFFFF", name="Arial", size=11)
def _bf(): return Font(bold=True, name="Arial", size=10)
def _rf(): return Font(name="Arial", size=10)
def _c():  return Alignment(horizontal="center", vertical="center")
def _l():  return Alignment(horizontal="left",   vertical="center")

def build_xlsx(results, title, subtitle):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    for letter, w in zip(COLS, WIDTHS):
        ws.column_dimensions[letter].width = w
    last = COLS[-1]

    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, name="Arial", size=13)
    ws["A1"].alignment = _c(); ws.row_dimensions[1].height = 24

    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="555555")
    ws["A2"].alignment = _c(); ws.row_dimensions[2].height = 16

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _hf(); c.fill = PatternFill("solid", fgColor=HDR_CLR)
        c.alignment = _c(); c.border = BDR
    ws.row_dimensions[3].height = 18

    row = 4
    s_start, s_end, cur = {}, {}, None

    for res in results:
        scen = res["scenario"]
        if scen != cur:
            if cur: s_end[cur] = row - 1
            cur = scen; s_start[scen] = row
            gf = PatternFill("solid", fgColor=SCEN_CLR[scen])
            ws.merge_cells(f"A{row}:{last}{row}")
            ws[f"A{row}"] = f"— {scen.upper()} TRAFFIC SCENARIOS —"
            ws[f"A{row}"].font = Font(bold=True, name="Arial", size=10, color="333333")
            ws[f"A{row}"].fill = gf; ws[f"A{row}"].alignment = _c()
            ws[f"A{row}"].border = BDR; ws.row_dimensions[row].height = 15; row += 1

        sf = PatternFill("solid", fgColor=SCEN_CLR[scen])
        vals = [res["label"], res["mean_veh_tl"], res["mean_ped_tl"],
                res["queue"], res["throughput"], res["north_queue"], res["south_queue"]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = sf; c.border = BDR
            c.alignment = _l() if col == 1 else _c()
            c.font = _rf()
            if col > 1: c.number_format = "0.00"
        ws.row_dimensions[row].height = 14; row += 1

    if cur: s_end[cur] = row - 1

    row += 1
    ws.merge_cells(f"A{row}:{last}{row}")
    ws[f"A{row}"] = "SCENARIO SUMMARIES"
    ws[f"A{row}"].font = Font(bold=True, name="Arial", size=11, color=HDR_CLR)
    ws[f"A{row}"].alignment = _c(); ws.row_dimensions[row].height = 18; row += 1

    for col, h in enumerate(["Scenario"] + HEADERS[1:], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _hf(); c.fill = PatternFill("solid", fgColor=HDR_CLR)
        c.alignment = _c(); c.border = BDR
    ws.row_dimensions[row].height = 16; row += 1

    for scen in ["normal", "slow", "jam"]:
        s, e = s_start.get(scen), s_end.get(scen)
        if not s or not e: continue
        fill = PatternFill("solid", fgColor=SCEN_CLR[scen])
        c = ws.cell(row=row, column=1, value=f"{scen.capitalize()} Traffic")
        c.font = _bf(); c.fill = fill; c.alignment = _l(); c.border = BDR
        for col, letter in enumerate(["B", "C", "D", "E", "F", "G"], 2):
            c = ws.cell(row=row, column=col, value=f"=AVERAGE({letter}{s+1}:{letter}{e})")
            c.font = _bf(); c.fill = fill; c.alignment = _c()
            c.border = BDR; c.number_format = "0.00"
        ws.row_dimensions[row].height = 14; row += 1

    row += 1
    of = PatternFill("solid", fgColor="D5E8D4")
    ws.merge_cells(f"A{row}:{last}{row}")
    ws[f"A{row}"] = "OVERALL MEAN (all 30 runs)"
    ws[f"A{row}"].font = Font(bold=True, name="Arial", size=10, color="1B5E20")
    ws[f"A{row}"].fill = of; ws[f"A{row}"].alignment = _c()
    ws[f"A{row}"].border = BDR; ws.row_dimensions[row].height = 16; row += 1

    all_rows = []
    for scen in ["normal", "slow", "jam"]:
        s, e = s_start.get(scen), s_end.get(scen)
        if s and e: all_rows.extend(range(s + 1, e + 1))

    for col, letter in enumerate(COLS, 1):
        c = ws.cell(row=row, column=col)
        c.fill = of; c.border = BDR; c.alignment = _c()
        if col == 1:
            c.value = "All Scenarios"; c.font = _bf()
        else:
            refs = ",".join(f"{letter}{r}" for r in all_rows)
            c.value = f"=AVERAGE({refs})"
            c.font = _bf(); c.number_format = "0.00"
    ws.row_dimensions[row].height = 14
    ws.freeze_panes = "A4"
    return wb

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    with open(DEBUG_LOG, 'w') as f:
        f.write("A2C Batch Test Debug Log\n")
        f.write(f"Action Space: {ACTION_SPACE}\n")
        f.write(f"Neutral action index: 5 (adjustment = 0)\n")
        f.write("="*70 + "\n")

    results = []
    for i, test in enumerate(TEST_PLAN, 1):
        print(f"\n[{i}/{len(TEST_PLAN)}] {test['label']}")
        if not os.path.isfile(test["rou_file"]):
            print(f"  SKIP — route file not found: {test['rou_file']}")
            results.append({"label": test["label"], "scenario": test["scenario"],
                             "mean_veh_tl": None, "mean_ped_tl": None,
                             "queue": None, "throughput": None,
                             "north_queue": None, "south_queue": None})
            continue
        try:
            vt, pt, ql, tp, nq, sq = run_simulation(test)
            results.append({"label": test["label"], "scenario": test["scenario"],
                             "mean_veh_tl": round(vt, 4), "mean_ped_tl": round(pt, 4),
                             "queue": round(ql, 4), "throughput": round(tp, 4),
                             "north_queue": round(nq, 4), "south_queue": round(sq, 4)})
        except Exception as e:
            print(f"  ERROR: {e}")
            import traceback
            traceback.print_exc()
            results.append({"label": test["label"], "scenario": test["scenario"],
                             "mean_veh_tl": None, "mean_ped_tl": None,
                             "queue": None, "throughput": None,
                             "north_queue": None, "south_queue": None})

    wb = build_xlsx(
        results,
        "Balibago Signalized Pedestrian — A2C Batch Test Results",
        "30 Runs  |  10 Normal  •  10 Slow  •  10 Jam  |  A2C (integrated loop)"
    )
    wb.save(XLSX_OUT)
    print(f"\n{'='*70}\n  Saved Excel: {XLSX_OUT}\n  Debug log:   {DEBUG_LOG}\n{'='*70}")

if __name__ == "__main__":
    main()
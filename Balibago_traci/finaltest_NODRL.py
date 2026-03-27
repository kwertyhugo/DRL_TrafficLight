"""
run_Balibago_SP_NoDRL_batch_test.py

Runs all 30 Balibago demand_test route files through the NoDRL SIGNALIZED
simulation and writes: Balibago_traci/batch_results/SP_NoDRL_Results.xlsx

Metrics per run (overall + North/South split):
  - Mean Veh. Time Loss (s)    — <tripinfo timeLoss>
  - Mean Ped. Time Loss (s)    — <personinfo timeLoss>
  - Mean Queue Length (m)      — avg jam across 13 detectors, every 60 s
  - Throughput (veh/hr)        — vehicles cleared, normalised to per-hour
  - Mean North Queue (m)       — avg jam e2_0–e2_7
  - Mean South Queue (m)       — avg jam e2_8–e2_12

Fixed signal timing matches SignalizedPedestrianNODRL.py exactly:
  North (4902876117): 8-phase, green = {0:45, 2:130, 4:30, 6:90}, yellow=5
  South (12188714):   8-phase, green = {0:25, 2:30,  4:40, 6:45}, yellow=5
"""

import os
import sys
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

if 'SUMO_HOME' in os.environ:
    sys.path.append(os.path.join(os.environ['SUMO_HOME'], 'tools'))
else:
    sys.exit("Please declare environment variable 'SUMO_HOME'")

import traci

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
SUMOCFG        = 'Balibago_traci/signalizedPed.sumocfg'
DEMAND_DIR     = 'Balibago_traci/demand_test'
OUTPUT_DIR     = 'Balibago_traci/batch_results/SP_NoDRL'
XLSX_OUT       = 'Balibago_traci/batch_results/SP_NoDRL_Results.xlsx'
STEP_LENGTH    = 0.1
MAX_STEPS      = 576000
METRIC_STEPS   = int(60 / STEP_LENGTH)

DETECTOR_IDS   = [f"e2_{i}" for i in range(13)]
DETECTOR_COUNT = 13
NORTH_IDS      = [f"e2_{i}" for i in range(8)]    # e2_0–e2_7
SOUTH_IDS      = [f"e2_{i}" for i in range(8,13)] # e2_8–e2_12

# Fixed signal timing — exact match to SignalizedPedestrianNODRL.py
NORTH_TL          = "4902876117"
NORTH_GREEN       = {0: 45, 2: 130, 4: 30, 6: 90}
NORTH_YELLOW      = 5
NORTH_PHASES      = 8
SOUTH_TL          = "12188714"
SOUTH_GREEN       = {0: 25, 2: 30, 4: 40, 6: 45}
SOUTH_YELLOW      = 5
SOUTH_PHASES      = 8

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(os.path.dirname(XLSX_OUT), exist_ok=True)

# ---------------------------------------------------------------------------
# Test plan
# ---------------------------------------------------------------------------
TEST_PLAN = []
for scenario in ["normal", "slow", "jam"]:
    for i in range(1, 11):
        TEST_PLAN.append({
            "label":     f"{scenario.capitalize()} {i}",
            "scenario":  scenario,
            "rou_file":  os.path.join(DEMAND_DIR, f"flows_{scenario}_traffic_{i:02d}.rou.xml"),
            "trips_out": os.path.join(OUTPUT_DIR, f"SP_NoDRL_trips_{scenario}_{i:02d}.xml"),
            "stats_out": os.path.join(OUTPUT_DIR, f"SP_NoDRL_stats_{scenario}_{i:02d}.xml"),
        })

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def subscribe_detectors():
    for det in DETECTOR_IDS:
        traci.lanearea.subscribe(det, [
            traci.constants.JAM_LENGTH_METERS,
            traci.constants.VAR_INTERVAL_NUMBER,
        ])


def apply_fixed_timing(n_phase, n_dur, s_phase, s_dur):
    n_dur -= STEP_LENGTH
    if n_dur <= 0:
        n_phase = (n_phase + 1) % NORTH_PHASES
        traci.trafficlight.setPhase(NORTH_TL, n_phase)
        n_dur = NORTH_YELLOW if n_phase % 2 == 1 else NORTH_GREEN.get(n_phase, 30)
        traci.trafficlight.setPhaseDuration(NORTH_TL, n_dur)
    s_dur -= STEP_LENGTH
    if s_dur <= 0:
        s_phase = (s_phase + 1) % SOUTH_PHASES
        traci.trafficlight.setPhase(SOUTH_TL, s_phase)
        s_dur = SOUTH_YELLOW if s_phase % 2 == 1 else SOUTH_GREEN.get(s_phase, 30)
        traci.trafficlight.setPhaseDuration(SOUTH_TL, s_dur)
    return n_phase, n_dur, s_phase, s_dur


def parse_tripinfo(path):
    try:
        root = ET.parse(path).getroot()
    except (FileNotFoundError, ET.ParseError):
        return 0.0, 0.0
    veh, ped = [], []
    for tag in root:
        tl = tag.get("timeLoss")
        if tl is None: continue
        (veh if tag.tag == "tripinfo" else ped).append(float(tl))
    return (sum(veh)/len(veh) if veh else 0.0,
            sum(ped)/len(ped) if ped else 0.0)


def run_simulation(test):
    sumo_cmd = [
        'sumo', '-c', SUMOCFG,
        '--route-files',        test["rou_file"],
        '--step-length',        str(STEP_LENGTH),
        '--delay',              '0',
        '--lateral-resolution', '0.1',
        '--statistic-output',   test["stats_out"],
        '--tripinfo-output',    test["trips_out"],
        '--no-warnings',        'true',
    ]
    print(f"\n{'='*65}\n  Running: {test['label']}\n{'='*65}")
    traci.start(sumo_cmd)
    subscribe_detectors()

    n_phase, n_dur = 0, NORTH_GREEN[0]
    s_phase, s_dur = 0, SOUTH_GREEN[0]
    step = 0
    jam_total = north_jam_total = south_jam_total = 0.0
    tp_total  = 0
    obs       = 0

    while traci.simulation.getMinExpectedNumber() > 0 and step < MAX_STEPS:
        n_phase, n_dur, s_phase, s_dur = apply_fixed_timing(n_phase, n_dur, s_phase, s_dur)

        if step % METRIC_STEPS == 0:
            obs += 1
            jam = north_jam = south_jam = tp = 0.0
            for det in DETECTOR_IDS:
                r = traci.lanearea.getSubscriptionResults(det)
                if r:
                    jl = r.get(traci.constants.JAM_LENGTH_METERS, 0)
                    tp += r.get(traci.constants.VAR_INTERVAL_NUMBER, 0)
                    jam += jl
                    if det in NORTH_IDS:
                        north_jam += jl
                    else:
                        south_jam += jl
            jam_total       += jam / DETECTOR_COUNT
            north_jam_total += north_jam / len(NORTH_IDS)
            south_jam_total += south_jam / len(SOUTH_IDS)
            tp_total        += tp

        traci.simulationStep()
        step += 1

    sim_s = step * STEP_LENGTH
    traci.close()

    mean_queue    = jam_total       / obs if obs else 0.0
    mean_n_queue  = north_jam_total / obs if obs else 0.0
    mean_s_queue  = south_jam_total / obs if obs else 0.0
    throughput_hr = tp_total / (sim_s / 3600.0) if sim_s > 0 else 0.0
    mean_veh_tl, mean_ped_tl = parse_tripinfo(test["trips_out"])

    print(f"  Queue:{mean_queue:.2f}m | N:{mean_n_queue:.2f} S:{mean_s_queue:.2f} | "
          f"VehTL:{mean_veh_tl:.2f}s | PedTL:{mean_ped_tl:.2f}s | TP:{throughput_hr:.1f}veh/hr")
    return mean_veh_tl, mean_ped_tl, mean_queue, throughput_hr, mean_n_queue, mean_s_queue

# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

SCEN_CLR = {"normal":"D6E4F7","slow":"FFF2CC","jam":"FCE4D6"}
HDR_CLR  = "2E4057"
thin = Side(style="thin", color="AAAAAA")
BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)
COLS    = ["A","B","C","D","E","F","G"]
WIDTHS  = [18, 22, 22, 22, 20, 20, 20]
HEADERS = ["Run",
           "Mean Veh. Time Loss (s)", "Mean Ped. Time Loss (s)",
           "Mean Queue Length (m)",   "Throughput (veh/hr)",
           "Mean North Queue (m)",    "Mean South Queue (m)"]

def _hf(): return Font(bold=True, color="FFFFFF", name="Arial", size=11)
def _bf(): return Font(bold=True, name="Arial", size=10)
def _rf(): return Font(name="Arial", size=10)
def _c():  return Alignment(horizontal="center", vertical="center")
def _l():  return Alignment(horizontal="left",   vertical="center")

def build_xlsx(results, title, subtitle):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    for letter, w in zip(COLS, WIDTHS):
        ws.column_dimensions[letter].width = w
    last = COLS[-1]

    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, name="Arial", size=13)
    ws["A1"].alignment = _c(); ws.row_dimensions[1].height = 24

    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="555555")
    ws["A2"].alignment = _c(); ws.row_dimensions[2].height = 16

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _hf(); c.fill = PatternFill("solid", fgColor=HDR_CLR)
        c.alignment = _c(); c.border = BDR
    ws.row_dimensions[3].height = 18

    row = 4
    s_start, s_end, cur = {}, {}, None

    for res in results:
        scen = res["scenario"]
        if scen != cur:
            if cur: s_end[cur] = row - 1
            cur = scen; s_start[scen] = row
            gf = PatternFill("solid", fgColor=SCEN_CLR[scen])
            ws.merge_cells(f"A{row}:{last}{row}")
            ws[f"A{row}"] = f"— {scen.upper()} TRAFFIC SCENARIOS —"
            ws[f"A{row}"].font = Font(bold=True, name="Arial", size=10, color="333333")
            ws[f"A{row}"].fill = gf; ws[f"A{row}"].alignment = _c()
            ws[f"A{row}"].border = BDR; ws.row_dimensions[row].height = 15; row += 1

        sf = PatternFill("solid", fgColor=SCEN_CLR[scen])
        vals = [res["label"], res["mean_veh_tl"], res["mean_ped_tl"],
                res["queue"], res["throughput"], res["north_queue"], res["south_queue"]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = sf; c.border = BDR
            c.alignment = _l() if col == 1 else _c()
            c.font = _rf()
            if col > 1: c.number_format = "0.00"
        ws.row_dimensions[row].height = 14; row += 1

    if cur: s_end[cur] = row - 1

    # Summaries
    row += 1
    ws.merge_cells(f"A{row}:{last}{row}")
    ws[f"A{row}"] = "SCENARIO SUMMARIES"
    ws[f"A{row}"].font = Font(bold=True, name="Arial", size=11, color=HDR_CLR)
    ws[f"A{row}"].alignment = _c(); ws.row_dimensions[row].height = 18; row += 1

    for col, h in enumerate(["Scenario"] + HEADERS[1:], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _hf(); c.fill = PatternFill("solid", fgColor=HDR_CLR)
        c.alignment = _c(); c.border = BDR
    ws.row_dimensions[row].height = 16; row += 1

    for scen in ["normal","slow","jam"]:
        s, e = s_start.get(scen), s_end.get(scen)
        if not s or not e: continue
        fill = PatternFill("solid", fgColor=SCEN_CLR[scen])
        c = ws.cell(row=row, column=1, value=f"{scen.capitalize()} Traffic")
        c.font = _bf(); c.fill = fill; c.alignment = _l(); c.border = BDR
        for col, letter in enumerate(["B","C","D","E","F","G"], 2):
            c = ws.cell(row=row, column=col,
                        value=f"=AVERAGE({letter}{s+1}:{letter}{e})")
            c.font = _bf(); c.fill = fill; c.alignment = _c()
            c.border = BDR; c.number_format = "0.00"
        ws.row_dimensions[row].height = 14; row += 1

    row += 1
    of = PatternFill("solid", fgColor="D5E8D4")
    ws.merge_cells(f"A{row}:{last}{row}")
    ws[f"A{row}"] = "OVERALL MEAN (all 30 runs)"
    ws[f"A{row}"].font = Font(bold=True, name="Arial", size=10, color="1B5E20")
    ws[f"A{row}"].fill = of; ws[f"A{row}"].alignment = _c()
    ws[f"A{row}"].border = BDR; ws.row_dimensions[row].height = 16; row += 1

    all_rows = []
    for scen in ["normal","slow","jam"]:
        s, e = s_start.get(scen), s_end.get(scen)
        if s and e: all_rows.extend(range(s+1, e+1))

    for col, letter in enumerate(COLS, 1):
        c = ws.cell(row=row, column=col)
        c.fill = of; c.border = BDR; c.alignment = _c()
        if col == 1:
            c.value = "All Scenarios"; c.font = _bf()
        else:
            refs = ",".join(f"{letter}{r}" for r in all_rows)
            c.value = f"=AVERAGE({refs})"
            c.font = _bf(); c.number_format = "0.00"
    ws.row_dimensions[row].height = 14
    ws.freeze_panes = "A4"
    return wb

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    results = []
    for i, test in enumerate(TEST_PLAN, 1):
        print(f"\n[{i}/{len(TEST_PLAN)}] {test['label']}")
        if not os.path.isfile(test["rou_file"]):
            print(f"  SKIP — not found: {test['rou_file']}")
            results.append({"label":test["label"],"scenario":test["scenario"],
                             "mean_veh_tl":None,"mean_ped_tl":None,"queue":None,
                             "throughput":None,"north_queue":None,"south_queue":None})
            continue
        try:
            vt, pt, ql, tp, nq, sq = run_simulation(test)
            results.append({"label":test["label"],"scenario":test["scenario"],
                             "mean_veh_tl":round(vt,4),"mean_ped_tl":round(pt,4),
                             "queue":round(ql,4),"throughput":round(tp,4),
                             "north_queue":round(nq,4),"south_queue":round(sq,4)})
        except Exception as e:
            print(f"  ERROR: {e}")
            results.append({"label":test["label"],"scenario":test["scenario"],
                             "mean_veh_tl":None,"mean_ped_tl":None,"queue":None,
                             "throughput":None,"north_queue":None,"south_queue":None})

    wb = build_xlsx(results,
                    "Balibago Signalized Pedestrian — NoDRL Batch Test Results",
                    "30 Runs  |  10 Normal  •  10 Slow  •  10 Jam  |  Fixed signal timing")
    wb.save(XLSX_OUT)
    print(f"\n{'='*65}\n  Saved: {XLSX_OUT}\n{'='*65}")

if __name__ == "__main__":
    main()
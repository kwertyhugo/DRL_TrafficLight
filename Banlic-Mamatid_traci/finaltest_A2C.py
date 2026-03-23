"""
run_BanlicMamatid_A2C_batch_test.py

Runs all 30 route files through the A2C SIGNALIZED simulation in test mode.
Output: Banlic-Mamatid_traci/batch_results/A2C_Results.xlsx

Matches TestSignalizedA2C.py exactly:
  - Loads .keras model file directly (not using agent class)
  - 11 discrete actions: (-25, -20, -15, -10, -5, 0, 5, 10, 15, 20, 25)
  - model.predict() for action selection
  - Phase transition inside apply_phase function
"""

import os
import sys
import xml.etree.ElementTree as ET
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from keras.utils import to_categorical
from keras.models import load_model

if 'SUMO_HOME' in os.environ:
    sys.path.append(os.path.join(os.environ['SUMO_HOME'], 'tools'))
else:
    sys.exit("Please declare environment variable 'SUMO_HOME'")

import traci

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
SUMOCFG     = 'Banlic-Mamatid_traci/signalizedPed.sumocfg'
DEMAND_DIR  = 'Banlic-Mamatid_traci/demand_test'
OUTPUT_DIR  = 'Banlic-Mamatid_traci/batch_results/A2C'
XLSX_OUT    = 'Banlic-Mamatid_traci/batch_results/A2C_Results.xlsx'
STEP_LENGTH = 0.1
MAX_STEPS   = 576000
METRIC_STEPS = int(60 / STEP_LENGTH)

DETECTOR_IDS    = ["e2_0", "e2_1", "e2_2", "e2_3", "e2_4", "e2_5"]
DETECTOR_COUNT  = len(DETECTOR_IDS)
JUNCTION_1      = "253768576"
JUNCTION_2      = "253499548"
TOTAL_PHASES    = 10
BASE_DURATIONS  = {0: 30, 2: 30, 4: 45, 6: 60, 8: 25}
YELLOW_DURATION = 5
ACTION_SPACE    = (-25, -20, -15, -10, -5, 0, 5, 10, 15, 20, 25)

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(os.path.dirname(XLSX_OUT), exist_ok=True)

# ---------------------------------------------------------------------------
# Load trained Keras model directly
# ---------------------------------------------------------------------------
MODEL_PATH = './Banlic-Mamatid_traci/models_A2C/Main_A2CAgent_Signalized.keras'

if not os.path.exists(MODEL_PATH):
    sys.exit(f"ERROR: A2C model not found at {MODEL_PATH}")

try:
    main_model = load_model(MODEL_PATH)
    print(f"\n{'='*65}\n  ✓ A2C model loaded from {MODEL_PATH}\n{'='*65}")
except Exception as e:
    sys.exit(f"ERROR: Could not load A2C model: {e}")

# ---------------------------------------------------------------------------
# Test plan
# ---------------------------------------------------------------------------
TEST_PLAN = []
for scenario in ["normal", "slow", "jam"]:
    for i in range(1, 11):
        TEST_PLAN.append({
            "label":     f"{scenario.capitalize()} {i}",
            "scenario":  scenario,
            "rou_file":  os.path.join(DEMAND_DIR, f"flows_{scenario}_traffic_{i:02d}.rou.xml"),
            "trips_out": os.path.join(OUTPUT_DIR, f"A2C_trips_{scenario}_{i:02d}.xml"),
            "stats_out": os.path.join(OUTPUT_DIR, f"A2C_stats_{scenario}_{i:02d}.xml"),
        })

# ---------------------------------------------------------------------------
# Subscriptions
# ---------------------------------------------------------------------------

def subscribe_detectors():
    ctx_vars = [traci.constants.VAR_TYPE, traci.constants.VAR_WAITING_TIME]
    met_vars = [traci.constants.JAM_LENGTH_METERS, traci.constants.VAR_INTERVAL_NUMBER]
    for det in DETECTOR_IDS:
        traci.lanearea.subscribeContext(det, traci.constants.CMD_GET_VEHICLE_VARIABLE, 3, ctx_vars)
        traci.lanearea.subscribe(det, met_vars)

# ---------------------------------------------------------------------------
# State helpers — match TestSignalizedA2C.py exactly
# ---------------------------------------------------------------------------

def weighted_waits(det_id):
    data = traci.lanearea.getContextSubscriptionResults(det_id)
    if not data:
        return 0
    weights = {"car": 1.0, "jeep": 1.5, "bus": 2.2, "truck": 2.5,
               "motorcycle": 0.3, "tricycle": 0.5}
    return sum(d.get(traci.constants.VAR_WAITING_TIME, 0)
               * weights.get(d.get(traci.constants.VAR_TYPE, "car"), 1.0)
               for d in data.values())

def intersection_queue():
    return [weighted_waits(det) for det in DETECTOR_IDS]

def predict_action(model, state):
    """Get greedy action from trained model (no exploration)."""
    state_batch = np.expand_dims(state, axis=0)
    action_probs = model.predict(state_batch, verbose=0)[0]
    return np.argmax(action_probs)

def trafficLight_phase(action_index, currentPhase, currentPhaseDuration):
    """
    Advance phase for both junctions — matches TestSignalizedA2C.py exactly.
    Phase increments INSIDE this function, then duration is applied to the NEW phase.
    """
    currentPhase = (currentPhase + 1) % TOTAL_PHASES

    traci.trafficlight.setPhase(JUNCTION_1, currentPhase)
    traci.trafficlight.setPhase(JUNCTION_2, currentPhase)

    if currentPhase % 2 == 1:  # yellow/transition
        currentPhaseDuration = YELLOW_DURATION
    else:  # green — apply action
        duration_adj = ACTION_SPACE[action_index]
        base = BASE_DURATIONS.get(currentPhase, 30)
        currentPhaseDuration = max(5, min(180, base + duration_adj))

    traci.trafficlight.setPhaseDuration(JUNCTION_1, currentPhaseDuration)
    traci.trafficlight.setPhaseDuration(JUNCTION_2, currentPhaseDuration)

    return currentPhase, currentPhaseDuration

# ---------------------------------------------------------------------------
# Tripinfo parser
# ---------------------------------------------------------------------------

def parse_tripinfo(path):
    try:
        root = ET.parse(path).getroot()
    except (FileNotFoundError, ET.ParseError):
        return 0.0, 0.0
    veh, ped = [], []
    for tag in root:
        tl = tag.get("timeLoss")
        if tl is None:
            continue
        (veh if tag.tag == "tripinfo" else ped).append(float(tl))
    return (sum(veh) / len(veh) if veh else 0.0,
            sum(ped) / len(ped) if ped else 0.0)

# ---------------------------------------------------------------------------
# Single run
# ---------------------------------------------------------------------------

def run_simulation(test):
    sumo_cmd = [
        'sumo', '-c', SUMOCFG,
        '--route-files',        test["rou_file"],
        '--step-length',        str(STEP_LENGTH),
        '--delay',              '0',
        '--lateral-resolution', '0.1',
        '--statistic-output',   test["stats_out"],
        '--tripinfo-output',    test["trips_out"],
        '--no-warnings',        'true',
    ]
    print(f"\n{'='*65}\n  Running: {test['label']}\n{'='*65}")
    traci.start(sumo_cmd)
    subscribe_detectors()

    currentPhase         = 0
    currentPhaseDuration = float(BASE_DURATIONS[0])
    prevAction           = None
    step                 = 0
    jam_total            = 0.0
    tp_total             = 0
    obs                  = 0

    while traci.simulation.getMinExpectedNumber() > 0 and step < MAX_STEPS:
        step += 1
        currentPhaseDuration -= STEP_LENGTH

        # Decision: only on green phases
        if currentPhaseDuration <= 0:
            if currentPhase % 2 == 0:  # green phase — ask the model
                queue = np.array(intersection_queue())
                norm_queue = queue / 1000.0
                phase_oh = to_categorical(currentPhase // 2, num_classes=5).flatten()
                obs_state = np.concatenate([norm_queue, phase_oh]).astype(np.float32)
                actionIndex = predict_action(main_model, obs_state)
                prevAction = actionIndex
            else:  # yellow — carry over last green action
                actionIndex = prevAction if prevAction is not None else 5

            # Apply phase transition (increments phase internally)
            currentPhase, currentPhaseDuration = trafficLight_phase(
                actionIndex, currentPhase, currentPhaseDuration)

        # Metrics every 60 s
        if step % METRIC_STEPS == 0:
            obs += 1
            jam = tp = 0.0
            for det in DETECTOR_IDS:
                r = traci.lanearea.getSubscriptionResults(det)
                if r:
                    jam += r.get(traci.constants.JAM_LENGTH_METERS, 0)
                    tp += r.get(traci.constants.VAR_INTERVAL_NUMBER, 0)
            jam_total += jam / DETECTOR_COUNT
            tp_total += tp

        traci.simulationStep()

    sim_s = step * STEP_LENGTH
    traci.close()

    mean_queue = jam_total / obs if obs else 0.0
    throughput_hr = tp_total / (sim_s / 3600.0) if sim_s > 0 else 0.0
    mean_veh_tl, mean_ped_tl = parse_tripinfo(test["trips_out"])

    print(f"  Queue:{mean_queue:.2f}m | VehTL:{mean_veh_tl:.2f}s | "
          f"PedTL:{mean_ped_tl:.2f}s | TP:{throughput_hr:.1f}veh/hr")
    return mean_veh_tl, mean_ped_tl, mean_queue, throughput_hr

# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

SCEN_CLR = {"normal": "D6E4F7", "slow": "FFF2CC", "jam": "FCE4D6"}
HDR_CLR = "2E4057"
thin = Side(style="thin", color="AAAAAA")
BDR = Border(left=thin, right=thin, top=thin, bottom=thin)
COLS = ["A", "B", "C", "D", "E"]
WIDTHS = [18, 22, 26, 22, 22]
HEADERS = ["Run", "Mean Veh. Time Loss (s)", "Mean Ped. Time Loss (s)",
           "Mean Queue Length (m)", "Throughput (veh/hr)"]


def _hf(): return Font(bold=True, color="FFFFFF", name="Arial", size=11)
def _bf(): return Font(bold=True, name="Arial", size=10)
def _rf(): return Font(name="Arial", size=10)
def _c(): return Alignment(horizontal="center", vertical="center")
def _l(): return Alignment(horizontal="left", vertical="center")


def build_xlsx(results, title, subtitle):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    for letter, w in zip(COLS, WIDTHS):
        ws.column_dimensions[letter].width = w
    last = COLS[-1]

    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, name="Arial", size=13)
    ws["A1"].alignment = _c()
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="555555")
    ws["A2"].alignment = _c()
    ws.row_dimensions[2].height = 16

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _hf()
        c.fill = PatternFill("solid", fgColor=HDR_CLR)
        c.alignment = _c()
        c.border = BDR
    ws.row_dimensions[3].height = 18

    row = 4
    s_start, s_end, cur = {}, {}, None

    for res in results:
        scen = res["scenario"]
        if scen != cur:
            if cur:
                s_end[cur] = row - 1
            cur = scen
            s_start[scen] = row
            gf = PatternFill("solid", fgColor=SCEN_CLR[scen])
            ws.merge_cells(f"A{row}:{last}{row}")
            ws[f"A{row}"] = f"— {scen.upper()} TRAFFIC SCENARIOS —"
            ws[f"A{row}"].font = Font(bold=True, name="Arial", size=10, color="333333")
            ws[f"A{row}"].fill = gf
            ws[f"A{row}"].alignment = _c()
            ws[f"A{row}"].border = BDR
            ws.row_dimensions[row].height = 15
            row += 1

        sf = PatternFill("solid", fgColor=SCEN_CLR[scen])
        for col, val in enumerate([res["label"], res["mean_veh_tl"], res["mean_ped_tl"],
                                    res["queue"], res["throughput"]], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = sf
            c.border = BDR
            c.alignment = _l() if col == 1 else _c()
            c.font = _rf()
            if col > 1:
                c.number_format = "0.00"
        ws.row_dimensions[row].height = 14
        row += 1

    if cur:
        s_end[cur] = row - 1

    row += 1
    ws.merge_cells(f"A{row}:{last}{row}")
    ws[f"A{row}"] = "SCENARIO SUMMARIES"
    ws[f"A{row}"].font = Font(bold=True, name="Arial", size=11, color=HDR_CLR)
    ws[f"A{row}"].alignment = _c()
    ws.row_dimensions[row].height = 18
    row += 1

    for col, h in enumerate(["Scenario"] + HEADERS[1:], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _hf()
        c.fill = PatternFill("solid", fgColor=HDR_CLR)
        c.alignment = _c()
        c.border = BDR
    ws.row_dimensions[row].height = 16
    row += 1

    for scen in ["normal", "slow", "jam"]:
        s, e = s_start.get(scen), s_end.get(scen)
        if not s or not e:
            continue
        fill = PatternFill("solid", fgColor=SCEN_CLR[scen])
        c = ws.cell(row=row, column=1, value=f"{scen.capitalize()} Traffic")
        c.font = _bf()
        c.fill = fill
        c.alignment = _l()
        c.border = BDR
        for col, letter in enumerate(["B", "C", "D", "E"], 2):
            c = ws.cell(row=row, column=col,
                        value=f"=AVERAGE({letter}{s+1}:{letter}{e})")
            c.font = _bf()
            c.fill = fill
            c.alignment = _c()
            c.border = BDR
            c.number_format = "0.00"
        ws.row_dimensions[row].height = 14
        row += 1

    row += 1
    of = PatternFill("solid", fgColor="D5E8D4")
    ws.merge_cells(f"A{row}:{last}{row}")
    ws[f"A{row}"] = "OVERALL MEAN (all 30 runs)"
    ws[f"A{row}"].font = Font(bold=True, name="Arial", size=10, color="1B5E20")
    ws[f"A{row}"].fill = of
    ws[f"A{row}"].alignment = _c()
    ws[f"A{row}"].border = BDR
    ws.row_dimensions[row].height = 16
    row += 1

    all_rows = []
    for scen in ["normal", "slow", "jam"]:
        s, e = s_start.get(scen), s_end.get(scen)
        if s and e:
            all_rows.extend(range(s + 1, e + 1))

    for col, letter in enumerate(COLS, 1):
        c = ws.cell(row=row, column=col)
        c.fill = of
        c.border = BDR
        c.alignment = _c()
        if col == 1:
            c.value = "All Scenarios"
            c.font = _bf()
        else:
            refs = ",".join(f"{letter}{r}" for r in all_rows)
            c.value = f"=AVERAGE({refs})"
            c.font = _bf()
            c.number_format = "0.00"
    ws.row_dimensions[row].height = 14
    ws.freeze_panes = "A4"
    return wb


def main():
    results = []
    for i, test in enumerate(TEST_PLAN, 1):
        print(f"\n[{i}/{len(TEST_PLAN)}] {test['label']}")
        if not os.path.isfile(test["rou_file"]):
            print(f"  SKIP — not found: {test['rou_file']}")
            results.append({"label": test["label"], "scenario": test["scenario"],
                            "mean_veh_tl": None, "mean_ped_tl": None,
                            "queue": None, "throughput": None})
            continue
        try:
            vt, pt, ql, tp = run_simulation(test)
            results.append({"label": test["label"], "scenario": test["scenario"],
                            "mean_veh_tl": round(vt, 4), "mean_ped_tl": round(pt, 4),
                            "queue": round(ql, 4), "throughput": round(tp, 4)})
        except Exception as e:
            print(f"  ERROR: {e}")
            import traceback
            traceback.print_exc()
            results.append({"label": test["label"], "scenario": test["scenario"],
                            "mean_veh_tl": None, "mean_ped_tl": None,
                            "queue": None, "throughput": None})

    wb = build_xlsx(results,
                    "Banlic-Mamatid Signalized — A2C Batch Test Results",
                    "30 Runs  |  10 Normal  •  10 Slow  •  10 Jam  |  A2C trained model (test mode)")
    wb.save(XLSX_OUT)
    print(f"\n{'='*65}\n  Saved: {XLSX_OUT}\n{'='*65}")


if __name__ == "__main__":
    main()
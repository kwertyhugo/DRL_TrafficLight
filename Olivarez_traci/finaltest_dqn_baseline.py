"""
run_BP_DQN_batch_test.py

Runs all 30 route files through the DQN BASELINE simulation in test mode.
Output: Olivarez_traci\batch_results\BP_DQN_Results.xlsx

Metrics per run:
  - Mean Vehicle Time Loss (s)     — mean timeLoss from <tripinfo> tags
  - Mean Pedestrian Time Loss (s)  — mean timeLoss from <personinfo> tags
  - Mean Queue Length (m)          — avg jam across 7 detectors, every 6 s
  - Throughput (veh/hr)            — vehicles cleared, normalised to per-hour

Run alongside run_SP_DQN_batch_test.py in a separate terminal.
"""

import os
import sys
import xml.etree.ElementTree as ET
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

if 'SUMO_HOME' in os.environ:
    sys.path.append(os.path.join(os.environ['SUMO_HOME'], 'tools'))
else:
    sys.exit("Please declare environment variable 'SUMO_HOME'")

import traci
from keras.utils import to_categorical
from models.DQN import DQNAgent as dqn

# ---------------------------------------------------------------------------
# Config — mirrors BaselinePedestrianDQN test mode exactly
# ---------------------------------------------------------------------------
SUMOCFG        = r'Olivarez_traci\baselinePed.sumocfg'
DEMAND_DIR     = r'Olivarez_traci\demand_test'
OUTPUT_DIR     = r'Olivarez_traci\batch_results\BP_DQN'
XLSX_OUT       = r'Olivarez_traci\batch_results\BP_DQN_Results.xlsx'
STEP_LENGTH    = 0.05
DETECTOR_IDS   = ["e2_4","e2_5","e2_6","e2_7","e2_8","e2_9","e2_10"]
DETECTOR_COUNT = len(DETECTOR_IDS)
METRIC_STEPS   = int(6 / STEP_LENGTH)           # every 6 s — matches original BP DQN
ACTION_SPACE   = (-25,-20,-15,-10,-5,0,5,10,15,20,25)

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(os.path.dirname(XLSX_OUT), exist_ok=True)

# ---------------------------------------------------------------------------
# Load trained agent
# ---------------------------------------------------------------------------
agent = dqn(state_size=13, action_size=11, memory_size=2000, gamma=0.95,
            epsilon=1.0, epsilon_decay_rate=0.995, epsilon_min=0,
            learning_rate=0.00005, target_update_freq=100,
            name='ReLU_DQNAgent_Baseline')
agent.load()

print(f"\n{'='*65}\n  BP DQN model loaded.\n{'='*65}")

# ---------------------------------------------------------------------------
# Test plan
# ---------------------------------------------------------------------------
TEST_PLAN = []
for scenario in ["normal", "slow", "jam"]:
    for i in range(1, 11):
        TEST_PLAN.append({
            "label":     f"{scenario.capitalize()} {i}",
            "scenario":  scenario,
            "rou_file":  os.path.join(DEMAND_DIR, f"flows_{scenario}_traffic_{i:02d}.rou.xml"),
            "trips_out": os.path.join(OUTPUT_DIR, f"BP_DQN_trips_{scenario}_{i:02d}.xml"),
            "stats_out": os.path.join(OUTPUT_DIR, f"BP_DQN_stats_{scenario}_{i:02d}.xml"),
        })

# ---------------------------------------------------------------------------
# Subscriptions
# ---------------------------------------------------------------------------

def subscribe_detectors():
    ctx_vars = [traci.constants.VAR_TYPE, traci.constants.VAR_WAITING_TIME]
    met_vars = [traci.constants.JAM_LENGTH_METERS, traci.constants.VAR_INTERVAL_NUMBER]
    for det in DETECTOR_IDS:
        traci.lanearea.subscribeContext(det, traci.constants.CMD_GET_VEHICLE_VARIABLE, 3, ctx_vars)
        traci.lanearea.subscribe(det, met_vars)

def subscribe_junctions():
    traci.junction.subscribeContext(
        "cluster_295373794_3477931123_7465167861",
        traci.constants.CMD_GET_PERSON_VARIABLE,
        10.0, [traci.constants.VAR_WAITING_TIME])

# ---------------------------------------------------------------------------
# State / phase helpers — exact match to original BP DQN
# ---------------------------------------------------------------------------

def weighted_waits(det_id):
    data = traci.lanearea.getContextSubscriptionResults(det_id)
    if not data: return 0
    wmap = {"car":1.0,"jeep":1.5,"bus":2.2,"truck":2.5,"motorcycle":0.3,"tricycle":0.5}
    return sum(d.get(traci.constants.VAR_WAITING_TIME, 0)
               * wmap.get(d.get(traci.constants.VAR_TYPE, "car"), 1.0)
               for d in data.values())

def get_state(phase):
    queues = [weighted_waits(d) for d in DETECTOR_IDS]
    sub = traci.junction.getContextSubscriptionResults("cluster_295373794_3477931123_7465167861")
    ped = sum(d.get(traci.constants.VAR_WAITING_TIME, 0) for d in sub.values()) if sub else 0
    queues.append(ped)
    norm_q    = np.array(queues) / 1000.0
    phase_enc = to_categorical(phase // 2, num_classes=5).flatten()
    return np.concatenate([norm_q, phase_enc]).astype(np.float32)

def apply_phase(action_idx, current_phase):
    next_phase = (current_phase + 1) % 10
    traci.trafficlight.setPhase("cluster_295373794_3477931123_7465167861", next_phase)
    if next_phase % 2 == 1:
        dur = 5
    else:
        base = 20 if next_phase == 4 else 30
        dur  = max(5, min(60, base + ACTION_SPACE[action_idx]))
    traci.trafficlight.setPhaseDuration("cluster_295373794_3477931123_7465167861", dur)
    return next_phase, float(dur)

# ---------------------------------------------------------------------------
# Tripinfo parser
# ---------------------------------------------------------------------------

def parse_tripinfo(path):
    try:
        root = ET.parse(path).getroot()
    except (FileNotFoundError, ET.ParseError):
        return 0.0, 0.0
    veh, ped = [], []
    for tag in root:
        tl = tag.get("timeLoss")
        if tl is None: continue
        (veh if tag.tag == "tripinfo" else ped).append(float(tl))
    return (sum(veh)/len(veh) if veh else 0.0,
            sum(ped)/len(ped) if ped else 0.0)

# ---------------------------------------------------------------------------
# Single run
# ---------------------------------------------------------------------------

def run_simulation(test):
    sumo_cmd = [
        'sumo', '-c', SUMOCFG,
        '--route-files',        test["rou_file"],
        '--step-length',        str(STEP_LENGTH),
        '--delay',              '0',
        '--lateral-resolution', '0.1',
        '--statistic-output',   test["stats_out"],
        '--tripinfo-output',    test["trips_out"],
        '--no-warnings',        'true',
    ]
    print(f"\n{'='*65}\n  Running: {test['label']}\n{'='*65}")
    traci.start(sumo_cmd)
    subscribe_detectors()
    subscribe_junctions()

    phase    = 0
    dur      = 30.0
    prev_act = 5
    step     = 0
    jam_total, tp_total, obs = 0.0, 0, 0

    while traci.simulation.getMinExpectedNumber() > 0:
        dur -= STEP_LENGTH

        if dur <= 0:
            if phase % 2 == 0:
                state    = get_state(phase)
                prev_act = agent.act(state)
            phase, dur = apply_phase(prev_act, phase)

        if step % METRIC_STEPS == 0:
            obs += 1
            for det in DETECTOR_IDS:
                r = traci.lanearea.getSubscriptionResults(det)
                if r:
                    jam_total += r.get(traci.constants.JAM_LENGTH_METERS, 0) / DETECTOR_COUNT
                    tp_total  += r.get(traci.constants.VAR_INTERVAL_NUMBER, 0)

        traci.simulationStep()
        step += 1

    sim_s = step * STEP_LENGTH
    traci.close()

    mean_queue    = jam_total / obs if obs else 0.0
    throughput_hr = tp_total / (sim_s / 3600.0) if sim_s > 0 else 0.0
    mean_veh_tl, mean_ped_tl = parse_tripinfo(test["trips_out"])

    print(f"  Queue:{mean_queue:.2f}m | VehTL:{mean_veh_tl:.2f}s | "
          f"PedTL:{mean_ped_tl:.2f}s | TP:{throughput_hr:.1f}veh/hr")
    return mean_veh_tl, mean_ped_tl, mean_queue, throughput_hr

# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

SCEN_CLR = {"normal":"D6E4F7","slow":"FFF2CC","jam":"FCE4D6"}
HDR_CLR  = "2E4057"
thin = Side(style="thin", color="AAAAAA")
BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)
COLS    = ["A","B","C","D","E"]
WIDTHS  = [18,22,26,22,22]
HEADERS = ["Run","Mean Veh. Time Loss (s)","Mean Ped. Time Loss (s)",
           "Mean Queue Length (m)","Throughput (veh/hr)"]

def _hf(): return Font(bold=True, color="FFFFFF", name="Arial", size=11)
def _bf(): return Font(bold=True, name="Arial", size=10)
def _rf(): return Font(name="Arial", size=10)
def _c():  return Alignment(horizontal="center", vertical="center")
def _l():  return Alignment(horizontal="left",   vertical="center")

def build_xlsx(results, title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    for letter, w in zip(COLS, WIDTHS):
        ws.column_dimensions[letter].width = w
    last = COLS[-1]

    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, name="Arial", size=13)
    ws["A1"].alignment = _c(); ws.row_dimensions[1].height = 24

    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = "30 Runs  |  10 Normal  •  10 Slow  •  10 Traffic Jam  |  DQN trained model (test mode)"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="555555")
    ws["A2"].alignment = _c(); ws.row_dimensions[2].height = 16

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _hf(); c.fill = PatternFill("solid", fgColor=HDR_CLR)
        c.alignment = _c(); c.border = BDR
    ws.row_dimensions[3].height = 18

    row = 4
    s_start, s_end, cur = {}, {}, None

    for res in results:
        scen = res["scenario"]
        if scen != cur:
            if cur: s_end[cur] = row - 1
            cur = scen; s_start[scen] = row
            gf = PatternFill("solid", fgColor=SCEN_CLR[scen])
            ws.merge_cells(f"A{row}:{last}{row}")
            ws[f"A{row}"] = f"— {scen.upper()} TRAFFIC SCENARIOS —"
            ws[f"A{row}"].font = Font(bold=True, name="Arial", size=10, color="333333")
            ws[f"A{row}"].fill = gf; ws[f"A{row}"].alignment = _c()
            ws[f"A{row}"].border = BDR; ws.row_dimensions[row].height = 15; row += 1

        sf = PatternFill("solid", fgColor=SCEN_CLR[scen])
        for col, val in enumerate([res["label"], res["mean_veh_tl"], res["mean_ped_tl"],
                                    res["queue"], res["throughput"]], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = sf; c.border = BDR
            c.alignment = _l() if col == 1 else _c()
            c.font = _rf()
            if col > 1: c.number_format = "0.00"
        ws.row_dimensions[row].height = 14; row += 1

    if cur: s_end[cur] = row - 1

    row += 1
    ws.merge_cells(f"A{row}:{last}{row}")
    ws[f"A{row}"] = "SCENARIO SUMMARIES"
    ws[f"A{row}"].font = Font(bold=True, name="Arial", size=11, color=HDR_CLR)
    ws[f"A{row}"].alignment = _c(); ws.row_dimensions[row].height = 18; row += 1

    for col, h in enumerate(["Scenario"] + HEADERS[1:], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _hf(); c.fill = PatternFill("solid", fgColor=HDR_CLR)
        c.alignment = _c(); c.border = BDR
    ws.row_dimensions[row].height = 16; row += 1

    for scen in ["normal","slow","jam"]:
        s, e = s_start.get(scen), s_end.get(scen)
        if not s or not e: continue
        fill = PatternFill("solid", fgColor=SCEN_CLR[scen])
        c = ws.cell(row=row, column=1, value=f"{scen.capitalize()} Traffic")
        c.font = _bf(); c.fill = fill; c.alignment = _l(); c.border = BDR
        for col, letter in enumerate(["B","C","D","E"], 2):
            c = ws.cell(row=row, column=col,
                        value=f"=AVERAGE({letter}{s+1}:{letter}{e})")
            c.font = _bf(); c.fill = fill; c.alignment = _c()
            c.border = BDR; c.number_format = "0.00"
        ws.row_dimensions[row].height = 14; row += 1

    row += 1
    of = PatternFill("solid", fgColor="D5E8D4")
    ws.merge_cells(f"A{row}:{last}{row}")
    ws[f"A{row}"] = "OVERALL MEAN (all 30 runs)"
    ws[f"A{row}"].font = Font(bold=True, name="Arial", size=10, color="1B5E20")
    ws[f"A{row}"].fill = of; ws[f"A{row}"].alignment = _c()
    ws[f"A{row}"].border = BDR; ws.row_dimensions[row].height = 16; row += 1

    all_rows = []
    for scen in ["normal","slow","jam"]:
        s, e = s_start.get(scen), s_end.get(scen)
        if s and e: all_rows.extend(range(s+1, e+1))

    for col, letter in enumerate(COLS, 1):
        c = ws.cell(row=row, column=col)
        c.fill = of; c.border = BDR; c.alignment = _c()
        if col == 1:
            c.value = "All Scenarios"; c.font = _bf()
        else:
            refs = ",".join(f"{letter}{r}" for r in all_rows)
            c.value = f"=AVERAGE({refs})"
            c.font = _bf(); c.number_format = "0.00"
    ws.row_dimensions[row].height = 14
    ws.freeze_panes = "A4"
    return wb

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    results = []
    for i, test in enumerate(TEST_PLAN, 1):
        print(f"\n[{i}/{len(TEST_PLAN)}] {test['label']}")
        if not os.path.isfile(test["rou_file"]):
            print(f"  SKIP — not found: {test['rou_file']}")
            results.append({"label":test["label"],"scenario":test["scenario"],
                             "mean_veh_tl":None,"mean_ped_tl":None,
                             "queue":None,"throughput":None})
            continue
        try:
            vt, pt, ql, tp = run_simulation(test)
            results.append({"label":test["label"],"scenario":test["scenario"],
                             "mean_veh_tl":round(vt,4),"mean_ped_tl":round(pt,4),
                             "queue":round(ql,4),"throughput":round(tp,4)})
        except Exception as e:
            print(f"  ERROR: {e}")
            results.append({"label":test["label"],"scenario":test["scenario"],
                             "mean_veh_tl":None,"mean_ped_tl":None,
                             "queue":None,"throughput":None})

    wb = build_xlsx(results, "Baseline Pedestrian — DQN Batch Test Results")
    wb.save(XLSX_OUT)
    print(f"\n{'='*65}\n  Saved: {XLSX_OUT}\n{'='*65}")

if __name__ == "__main__":
    main()